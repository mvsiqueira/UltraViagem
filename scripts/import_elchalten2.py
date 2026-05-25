import openpyxl, json, os, re, sys
from import_utils import cell_color, parse_text, build_merge_map, make_activity_fn, parse_expenses
sys.stdout.reconfigure(encoding='utf-8')

XLSX   = r'G:\Meu Drive\Viagens\2025-03 El Chaltén\Viagem El Chatén.xlsx'
FOLDER = r'G:\Meu Drive\Viagens\2025-03 El Chaltén'
PFX    = 'chalten-2025'

wb      = openpyxl.load_workbook(XLSX)
wb_data = openpyxl.load_workbook(XLSX, data_only=True)
ws      = wb['Roteiro']
merge_map = build_merge_map(ws)
act = make_activity_fn(ws, merge_map, first_col=3)

DATES = ["2025-03-22","2025-03-23","2025-03-24","2025-03-25",
         "2025-03-26","2025-03-27","2025-03-28","2025-03-29"]
SUMMARIES = ["IDA","El Chaltén","Fitz Roy","Cerro Torre",
             "Livre","Dos Glaciares","El Calafate","VOLTA"]

DAYS_DEF = [
    (3, 0, [(3,'a01','Transporte'),(8,'a02','Transporte'),(11,'a03','Hospedagem')]),
    (4, 1, [(4,'a04','Trilha'),(6,'a05','Refeição'),(7,'a06','Passeio'),(11,'a07','Hospedagem')]),
    (5, 2, [(3,'a08','Trilha'),(10,'a09','Hospedagem'),(11,'a10','Hospedagem')]),
    (6, 3, [(3,'a11','Trilha'),(11,'a12','Hospedagem')]),
    (7, 4, [(11,'a13','Hospedagem')]),
    (8, 5, [(3,'a14','Passeio'),(11,'a15','Hospedagem')]),
    (9, 6, [(3,'a16','Passeio'),(5,'a17','Transporte'),(7,'a18','Refeição'),(11,'a19','Hospedagem')]),
    (10,7, [(3,'a20','Transporte')]),
]

itinerary = []
for (row, di, cols) in DAYS_DEF:
    activities = []
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = act(f"{PFX}-{sid}", row, col, atype)
        if a: activities.append(a)
    itinerary.append({"id":f"{PFX}-d{di+1:02d}","date":DATES[di],
                      "title":f"Dia {di+1}","summary":SUMMARIES[di],"activities":activities})

BANK_DEF = [
    (12, 0, [(3,'b01','Trilha'),(5,'b02','Trilha'),(8,'b03','Trilha')]),
    (13, 1, [(3,'b04','Trilha')]),
    (14, 2, [(3,'b05','Trilha')]),
    (15, 3, [(3,'b06','Passeio')]),
]

bank_activities = []
for (row, br, cols) in BANK_DEF:
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = act(f"{PFX}-{sid}", row, col, atype, bankrow=br)
        if a: bank_activities.append(a)

ws_d = wb['Dicas']
links, seen_urls, lid = [], set(), 1
for row in ws_d.iter_rows(min_row=2):
    tv = row[0].value; uv = row[1].value if len(row)>1 else None
    if not tv or not uv: continue
    url = re.sub(r'\+[A-Z]\d+$','',str(uv).strip())
    if url in seen_urls: continue
    seen_urls.add(url)
    links.append({"id":f"{PFX}-l{lid:02d}","title":str(tv).strip(),"url":url})
    lid += 1

expenses = parse_expenses(wb['Gastos'], wb_data['Gastos'], PFX)

files = sorted([f for f in os.listdir(FOLDER)
                if not f.startswith('.') and os.path.isfile(os.path.join(FOLDER,f))])
attachments = [{"id":f"{PFX}-att{i+1:02d}","file":f} for i,f in enumerate(files)]

trip = {
    "schemaVersion":1,"id":PFX,"title":"2025-03 El Chaltén",
    "startDate":"2025-03-22","endDate":"2025-03-29",
    "baseCurrency":"BRL","people":2,"rateDecimalDigits":2,"myMapsUrl":"",
    "itinerarySlotsPerDay":9,
    "itineraryVersions":[{"id":f"{PFX}-v1","name":"Versão 1","bankRows":4,
                          "itinerary":itinerary,"bankActivities":bank_activities}],
    "activeVersionId":f"{PFX}-v1",
    "tasks":[],"links":links,"expenses":expenses,"currencyRates":[],"attachments":attachments
}

out = os.path.join(FOLDER,'trip.json')
with open(out,'w',encoding='utf-8') as f: json.dump(trip,f,ensure_ascii=False,indent=2)
print(f"OK: {out}")
print(f"  {len(itinerary)} dias | {sum(len(d['activities']) for d in itinerary)} ativ | {len(bank_activities)} banco | {len(links)} dicas | {len(expenses)} gastos | {len(attachments)} anexos")
print("\nAtividades:")
for d in itinerary:
    print(f"  {d['date']} {d['summary']}:")
    for a in d['activities']:
        det = f"  [{a.get('details','')[:30]}]" if a.get('details') else ''
        print(f"    [{a['startSlot']}+{a['durationSlots']}] {a['type']:12} {a['title']}{det}")
print("\nGastos:")
for e in expenses:
    print(f"  {e['id']}: {e['title'][:40]:40} {e['currency']} p={e['price']} ppl={e['people']} qty={e['quantity']} paid={e['paidAmount']}")
