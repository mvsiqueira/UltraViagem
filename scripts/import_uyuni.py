import openpyxl, json, os, re, sys
from import_utils import build_merge_map, make_activity_fn
sys.stdout.reconfigure(encoding='utf-8')

XLSX   = r'G:\Meu Drive\Viagens\2024-04 Salar de Uyuni\Viagem Salar de Uyuni.xlsx'
FOLDER = r'G:\Meu Drive\Viagens\2024-04 Salar de Uyuni'
PFX    = 'uyuni-2024'

wb      = openpyxl.load_workbook(XLSX)
wb_data = openpyxl.load_workbook(XLSX, data_only=True)
ws      = wb['Roteiro']
merge_map = build_merge_map(ws)
act = make_activity_fn(ws, merge_map, first_col=4)

DATES = ["2024-04-19","2024-04-20","2024-04-21","2024-04-22","2024-04-23",
         "2024-04-24","2024-04-25","2024-04-26","2024-04-27","2024-04-28"]
SUMMARIES = ["IDA","Descanso","Atacama 1","Atacama 2","Salar de Uyuni 1",
             "Salar de Uyuni 2","Salar de Uyuni 3","Salar de Uyuni 4","Vulcao Lascar","VOLTA"]

DAYS_DEF = [
    (3, 0, [(4,'a01','Transporte'),(11,'a02','Transporte'),(12,'a03','Hospedagem')]),
    (4, 1, [(4,'a04','Passeio'),(12,'a05','Hospedagem')]),
    (5, 2, [(4,'a06','Passeio'),(12,'a07','Hospedagem')]),
    (6, 3, [(4,'a08','Trilha'),(12,'a09','Hospedagem')]),
    (7, 4, [(4,'a10','Passeio'),(12,'a11','Hospedagem')]),
    (8, 5, [(4,'a12','Passeio'),(12,'a13','Hospedagem')]),
    (9, 6, [(4,'a14','Passeio'),(12,'a15','Hospedagem')]),
    (10,7, [(4,'a16','Passeio'),(8,'a17','Passeio'),(12,'a18','Hospedagem')]),
    (11,8, [(4,'a19','Trilha'),(12,'a20','Hospedagem')]),
    (12,9, [(4,'a21','Transporte'),(5,'a22','Transporte')]),
]

itinerary = []
for (row, di, cols) in DAYS_DEF:
    activities = []
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = act(f"{PFX}-{sid}", row, col, atype)
        if a: activities.append(a)
    itinerary.append({"id":f"{PFX}-d{di+1:02d}","date":DATES[di],
                      "title":f"Dia {di+1}","summary":SUMMARIES[di],"activities":activities})

BANK_DEF = [
    (14, 0, [(4,'b01','Passeio'),(8,'b02','Passeio')]),
]
bank_activities = []
for (row, br, cols) in BANK_DEF:
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = act(f"{PFX}-{sid}", row, col, atype, bankrow=br)
        if a: bank_activities.append(a)

ws_d = wb['Dicas']
links, seen_urls, lid = [], set(), 1
maps_url = ""
for row in ws_d.iter_rows(min_row=2):
    tv = row[0].value; uv = row[1].value if len(row)>1 else None
    if not tv or not uv: continue
    url = re.sub(r'\+[A-Z]\d+$','',str(uv).strip())
    title = str(tv).strip()
    if title == 'Mapa': maps_url = url; continue
    if url in seen_urls: continue
    seen_urls.add(url)
    links.append({"id":f"{PFX}-l{lid:02d}","title":title,"url":url}); lid += 1

def parse_gastos_uyuni(ws_g, ws_g_data, pfx):
    expenses = []; eid = 1
    for rr, rd in zip(ws_g.iter_rows(min_row=2), ws_g_data.iter_rows(min_row=2)):
        raw = list(rr); dat = list(rd)
        if str(raw[0].value or '').strip() not in (chr(9658), chr(9654)): continue
        title  = str(raw[1].value or '').strip()
        etype  = str(raw[2].value or '').strip() or None
        obs1   = str(raw[3].value or '').strip()
        obs2   = str(raw[4].value or '').strip() if len(raw)>4 else ''
        obs3   = str(raw[5].value or '').strip() if len(raw)>5 else ''
        obs4   = str(raw[6].value or '').strip() if len(raw)>6 else ''
        price_r = raw[8].value; price_d = dat[8].value
        price = float(price_r) if isinstance(price_r,(int,float)) else (float(price_d) if isinstance(price_d,(int,float)) else 0.0)
        taxes_r = raw[9].value
        taxes = float(taxes_r) if isinstance(taxes_r,(int,float)) else 0.0
        people = int(round(float(raw[10].value))) if isinstance(raw[10].value,(int,float)) else 1
        qty    = int(raw[11].value or 1) if raw[11].value else 1
        moeda_raw  = str(raw[13].value or '1').strip()
        moeda_calc = dat[13].value
        # Detect currency: named formula OR hardcoded rate
        currency, rate = 'BRL', 1.0
        if 'Dolar' in moeda_raw or chr(243) in moeda_raw:  # Dólar
            currency = 'USD'; rate = float(moeda_calc) if isinstance(moeda_calc,(int,float)) else 0.0
        elif 'Peso' in moeda_raw:
            currency = 'CLP'; rate = float(moeda_calc) if isinstance(moeda_calc,(int,float)) else 0.0
        elif moeda_raw == '1':
            currency = 'BRL'; rate = 1.0
        else:
            try:
                r = float(moeda_raw)
                if r > 2.0: currency, rate = 'USD', r
                elif r < 0.1: currency, rate = 'CLP', r
            except ValueError: pass
        paid_d = dat[15].value; paid_r = raw[15].value
        paid = float(paid_d) if isinstance(paid_d,(int,float)) else (float(paid_r) if isinstance(paid_r,(int,float)) else 0.0)
        if paid == 0 and price > 0:
            paid = round((price + taxes) * people * qty * rate, 2)
        if not title: title = obs1 if obs1 and obs1!='-' else (obs2 if obs2 and obs2!='-' else 'Item')
        notes = ' . '.join(p for p in [obs1,obs2,obs3,obs4] if p and p not in ('-','')) or None
        if etype == 'Hotel': etype = 'Hospedagem'
        if etype == 'Atividades': etype = 'Passeios'
        if not etype:
            tl = title.lower()
            if any(w in tl for w in ['hotel','pousada','lodge','hostel']): etype='Hospedagem'
            elif any(w in tl for w in ['traslado','transfer','voo','voo']): etype='Transporte'
            elif any(w in tl for w in ['ingresso','excursao','trekking','lagunas','salar','vulcao']): etype='Passeios'
        e = {"id":f"{pfx}-e{eid:02d}","isActive":True,"title":title}
        if etype: e["type"] = etype
        if notes: e["notes"] = notes
        e.update({"price":round(price,2),"taxes":round(taxes,2),"people":people,"quantity":qty,
                  "currency":currency,"exchangeRateToBase":round(rate,6),"paidAmount":round(paid,2)})
        expenses.append(e); eid += 1
    return expenses

expenses = parse_gastos_uyuni(wb['Gastos'], wb_data['Gastos'], PFX)

files = sorted([f for f in os.listdir(FOLDER) if not f.startswith('.') and os.path.isfile(os.path.join(FOLDER,f))])
attachments = [{"id":f"{PFX}-att{i+1:02d}","file":f} for i,f in enumerate(files)]

trip = {
    "schemaVersion":1,"id":PFX,"title":"2024-04 Salar de Uyuni",
    "startDate":"2024-04-19","endDate":"2024-04-28",
    "baseCurrency":"BRL","people":2,"rateDecimalDigits":2,"myMapsUrl":maps_url,
    "itinerarySlotsPerDay":9,
    "itineraryVersions":[{"id":f"{PFX}-v1","name":"Versao 1","bankRows":1,
                          "itinerary":itinerary,"bankActivities":bank_activities}],
    "activeVersionId":f"{PFX}-v1",
    "tasks":[],"links":links,"expenses":expenses,"currencyRates":[],"attachments":attachments
}

out = os.path.join(FOLDER,'trip.json')
with open(out,'w',encoding='utf-8') as f: json.dump(trip,f,ensure_ascii=False,indent=2)
print(f"OK: {out}")
print(f"  {len(itinerary)} dias | {sum(len(d['activities']) for d in itinerary)} ativ | {len(bank_activities)} banco | {len(links)} dicas | {len(expenses)} gastos | {len(attachments)} anexos")
