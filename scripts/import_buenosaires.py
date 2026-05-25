import openpyxl, json, os, re, sys
from import_utils import cell_color, parse_text, build_merge_map, make_activity_fn
sys.stdout.reconfigure(encoding='utf-8')

XLSX   = r'G:\Meu Drive\Viagens\2023-02 Buenos Aires\Viagem Buenos Aires1.xlsx'
FOLDER = r'G:\Meu Drive\Viagens\2023-02 Buenos Aires'
PFX    = 'buenosaires-2023'

wb      = openpyxl.load_workbook(XLSX)
wb_data = openpyxl.load_workbook(XLSX, data_only=True)
ws      = wb['Roteiro']
merge_map = build_merge_map(ws)
act = make_activity_fn(ws, merge_map, first_col=4)

DATES     = ["2023-02-19","2023-02-20","2023-02-21","2023-02-22","2023-02-23"]
SUMMARIES = ["IDA","Buenos Aires 1","Buenos Aires 2","Parque","VOLTA"]

DAYS_DEF = [
    (3, 0, [(4,'a01','Transporte'),(6,'a02','Refeição'),(7,'a03','Passeio'),(10,'a04','Hospedagem')]),
    (4, 1, [(4,'a05','Passeio'),(6,'a06','Refeição'),(7,'a07','Passeio'),(10,'a08','Hospedagem')]),
    (5, 2, [(4,'a09','Passeio'),(6,'a10','Refeição'),(7,'a11','Passeio'),(10,'a12','Hospedagem')]),
    (6, 3, [(6,'a13','Refeição'),(10,'a14','Hospedagem')]),
    (7, 4, [(4,'a15','Passeio'),(6,'a16','Refeição'),(8,'a17','Transporte')]),
]

itinerary = []
for (row, di, cols) in DAYS_DEF:
    activities = []
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = act(f"{PFX}-{sid}", row, col, atype)
        if a: activities.append(a)
    itinerary.append({"id":f"{PFX}-d{di+1:02d}","date":DATES[di],
                      "title":f"Dia {di+1}","summary":SUMMARIES[di],"activities":activities})

BANK_DEF = [
    (9,  0, [(4,'b01','Passeio'),(5,'b02','Passeio'),(6,'b03','Passeio'),(7,'b04','Passeio'),(8,'b05','Passeio')]),
    (10, 1, [(4,'b06','Passeio')]),
]
bank_activities = []
for (row, br, cols) in BANK_DEF:
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = act(f"{PFX}-{sid}", row, col, atype, bankrow=br)
        if a: bank_activities.append(a)

ws_d = wb['Dicas']
links, seen_urls, lid = [], set(), 1
maps_url = ""
for row in ws_d.iter_rows(min_row=2):
    tv = row[0].value; uv = row[1].value if len(row)>1 else None
    if not tv or not uv: continue
    url = re.sub(r'\+[A-Z]\d+$','',str(uv).strip())
    title = str(tv).strip()
    if title == 'Mapa': maps_url = url; continue
    if url in seen_urls: continue
    seen_urls.add(url)
    links.append({"id":f"{PFX}-l{lid:02d}","title":title,"url":url})
    lid += 1

# Gastos — BA: no taxes; price@8, pessoas@9, qty@10, moeda@12, pago@14
def parse_gastos_ba(ws_g, ws_g_data, pfx):
    expenses = []
    eid = 1
    rows_raw  = list(ws_g.iter_rows(min_row=2))
    rows_data = list(ws_g_data.iter_rows(min_row=2))
    moeda_map = {'ARS':'ARS','Dólar':'USD','Dolar':'USD'}
    for row_raw, row_dat in zip(rows_raw, rows_data):
        raw = list(row_raw); dat = list(row_dat)
        marker = str(raw[0].value or '').strip()
        if marker not in ('►','▶'): continue
        title   = str(raw[1].value or '').strip()
        etype   = str(raw[2].value or '').strip() or None
        empresa = str(raw[3].value or '').strip()
        obs2    = str(raw[4].value or '').strip()
        obs3    = str(raw[5].value or '').strip()
        price_r = raw[8].value; price_d = dat[8].value
        if isinstance(price_r,(int,float)): price = float(price_r)
        elif isinstance(price_d,(int,float)): price = float(price_d)
        else: price = 0.0
        ppl_raw = raw[9].value
        people = int(round(float(ppl_raw))) if isinstance(ppl_raw,(int,float)) and ppl_raw else 1
        qty    = int(raw[10].value or 1) if raw[10].value else 1
        moeda_raw  = str(raw[12].value or '1').strip()
        moeda_calc = dat[12].value
        currency, rate = 'BRL', 1.0
        for key, cur in moeda_map.items():
            if key in moeda_raw:
                currency = cur
                rate = float(moeda_calc) if isinstance(moeda_calc,(int,float)) else 0.0
                break
        paid_r = raw[14].value; paid_d = dat[14].value
        if isinstance(paid_d,(int,float)): paid = float(paid_d)
        elif isinstance(paid_r,(int,float)): paid = float(paid_r)
        else: paid = 0.0
        if not title:
            title = empresa if empresa and empresa != '-' else \
                    (obs2 if obs2 and obs2 != '-' else 'Item')
        note_parts = [p for p in [empresa,obs2,obs3] if p and p not in ('-','')]
        notes = ' · '.join(note_parts) if note_parts else None
        if etype == 'Hotel': etype = 'Hospedagem'
        if etype == 'Atividades': etype = 'Passeios'
        e = {"id":f"{pfx}-e{eid:02d}","isActive":True,"title":title}
        if etype:  e["type"]    = etype
        if notes:  e["notes"]   = notes
        e["price"]=round(price,2); e["taxes"]=0.0; e["people"]=people
        e["quantity"]=qty; e["currency"]=currency
        e["exchangeRateToBase"]=round(rate,6); e["paidAmount"]=round(paid,2)
        expenses.append(e); eid += 1
    return expenses

expenses = parse_gastos_ba(wb['Gastos'], wb_data['Gastos'], PFX)

files = sorted([f for f in os.listdir(FOLDER)
                if not f.startswith('.') and os.path.isfile(os.path.join(FOLDER,f))])
attachments = [{"id":f"{PFX}-att{i+1:02d}","file":f} for i,f in enumerate(files)]

trip = {
    "schemaVersion":1,"id":PFX,"title":"2023-02 Buenos Aires",
    "startDate":"2023-02-19","endDate":"2023-02-23",
    "baseCurrency":"BRL","people":5,"rateDecimalDigits":2,"myMapsUrl":maps_url,
    "itinerarySlotsPerDay":7,
    "itineraryVersions":[{"id":f"{PFX}-v1","name":"Versão 1","bankRows":2,
                          "itinerary":itinerary,"bankActivities":bank_activities}],
    "activeVersionId":f"{PFX}-v1",
    "tasks":[],"links":links,"expenses":expenses,"currencyRates":[],"attachments":attachments
}

out = os.path.join(FOLDER,'trip.json')
with open(out,'w',encoding='utf-8') as f: json.dump(trip,f,ensure_ascii=False,indent=2)
print(f"OK: {out}")
print(f"  {len(itinerary)} dias | {sum(len(d['activities']) for d in itinerary)} ativ | {len(bank_activities)} banco | {len(links)} dicas | {len(expenses)} gastos | {len(attachments)} anexos")
