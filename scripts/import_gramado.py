import openpyxl, json, os, re, sys
from import_utils import build_merge_map, make_activity_fn
sys.stdout.reconfigure(encoding='utf-8')

XLSX   = r'G:\Meu Drive\Viagens\2022-01 Gramado\Viagem Gramado.xlsx'
FOLDER = r'G:\Meu Drive\Viagens\2022-01 Gramado'
PFX    = 'gramado-2022'

wb      = openpyxl.load_workbook(XLSX)
wb_data = openpyxl.load_workbook(XLSX, data_only=True)
ws      = wb['Roteiro']
merge_map = build_merge_map(ws)
act = make_activity_fn(ws, merge_map, first_col=4)

DATES = ["2022-02-27","2022-02-28","2022-03-01","2022-03-02","2022-03-03","2022-03-04"]
SUMMARIES = ["Ida","Gramado","Gramado","Canela","Canela","Volta"]

DAYS_DEF = [
    (3, 0, [(4,'a01','Transporte'),(6,'a02','Transporte'),(7,'a03','Refeicao'),
            (8,'a04','Passeio'),(10,'a05','Hospedagem')]),
    (4, 1, [(4,'a06','Passeio'),(6,'a07','Refeicao'),(7,'a08','Passeio'),
            (8,'a09','Passeio'),(9,'a10','Passeio'),(10,'a11','Hospedagem')]),
    (5, 2, [(4,'a12','Passeio'),(6,'a13','Refeicao'),(7,'a14','Passeio'),
            (9,'a15','Passeio'),(10,'a16','Hospedagem')]),
    (6, 3, [(4,'a17','Passeio'),(6,'a18','Refeicao'),(7,'a19','Passeio'),
            (10,'a20','Hospedagem')]),
    (7, 4, [(4,'a21','Passeio'),(6,'a22','Refeicao'),(7,'a23','Passeio'),
            (10,'a24','Hospedagem')]),
    (8, 5, [(4,'a25','Passeio'),(5,'a26','Transporte'),(6,'a27','Refeicao'),
            (7,'a28','Transporte'),(8,'a29','Passeio'),(9,'a30','Transporte'),
            (10,'a31','Transporte')]),
]

itinerary = []
for (row, di, cols) in DAYS_DEF:
    activities = []
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = act(f"{PFX}-{sid}", row, col, atype)
        if a: activities.append(a)
    itinerary.append({"id":f"{PFX}-d{di+1:02d}","date":DATES[di],
                      "title":f"Dia {di+1}","summary":SUMMARIES[di],"activities":activities})

BANK_DEF = [
    (12, 0, [(4,'b01','Passeio'),(6,'b02','Passeio'),(7,'b03','Passeio'),(8,'b04','Passeio'),(9,'b05','Passeio')]),
    (13, 1, [(4,'b06','Refeicao'),(5,'b07','Passeio'),(6,'b08','Passeio'),(7,'b09','Passeio'),(8,'b10','Passeio'),(9,'b11','Passeio')]),
    (14, 2, [(4,'b12','Passeio'),(5,'b13','Passeio'),(6,'b14','Passeio'),(7,'b15','Passeio'),(8,'b16','Passeio'),(9,'b17','Passeio')]),
    (15, 3, [(4,'b18','Passeio'),(5,'b19','Passeio'),(6,'b20','Refeicao'),(7,'b21','Refeicao'),(8,'b22','Refeicao')]),
    (16, 4, [(4,'b23','Passeio'),(5,'b24','Passeio'),(6,'b25','Passeio'),(7,'b26','Passeio'),(8,'b27','Passeio'),(9,'b28','Passeio')]),
    (17, 5, [(4,'b29','Passeio'),(5,'b30','Passeio'),(6,'b31','Passeio'),(7,'b32','Passeio'),(8,'b33','Passeio'),(9,'b34','Passeio')]),
    (18, 6, [(4,'b35','Passeio'),(5,'b36','Passeio'),(6,'b37','Passeio'),(7,'b38','Passeio'),(8,'b39','Passeio'),(9,'b40','Passeio')]),
    (19, 7, [(4,'b41','Passeio'),(5,'b42','Passeio'),(6,'b43','Passeio'),(7,'b44','Passeio'),(8,'b45','Passeio'),(9,'b46','Passeio')]),
    (20, 8, [(4,'b47','Passeio'),(5,'b48','Passeio'),(6,'b49','Passeio'),(7,'b50','Passeio'),(8,'b51','Passeio'),(9,'b52','Refeicao')]),
    (21, 9, [(4,'b53','Passeio'),(5,'b54','Passeio'),(6,'b55','Passeio'),(7,'b56','Passeio'),(8,'b57','Passeio'),(9,'b58','Passeio')]),
    (22,10, [(4,'b59','Passeio')]),
    (23,11, [(4,'b60','Passeio')]),
    (24,12, [(4,'b61','Refeicao'),(5,'b62','Passeio')]),
    (25,13, [(4,'b63','Passeio')]),
]
bank_activities = []
for (row, br, cols) in BANK_DEF:
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = act(f"{PFX}-{sid}", row, col, atype, bankrow=br)
        if a: bank_activities.append(a)

ws_d = wb['Dicas']
links, seen_urls, lid = [], set(), 1
maps_url = ""
for row in ws_d.iter_rows(min_row=2):
    tv = row[0].value; uv = row[1].value if len(row)>1 else None
    if not tv or not uv: continue
    url = re.sub(r'\+[A-Z]\d+$','',str(uv).strip())
    title = str(tv).strip()
    if not url.startswith('http'): continue
    if title == 'Mapa': maps_url = url; continue
    if url in seen_urls: continue
    seen_urls.add(url)
    links.append({"id":f"{PFX}-l{lid:02d}","title":title,"url":url}); lid += 1

def parse_gastos_gramado(ws_g, ws_g_data, pfx):
    expenses = []; eid = 1
    for rr, rd in zip(ws_g.iter_rows(min_row=2), ws_g_data.iter_rows(min_row=2)):
        raw = list(rr); dat = list(rd)
        if str(raw[0].value or '').strip() not in (chr(9658), chr(9654), '>'): continue
        title = str(raw[1].value or '').strip()
        etype = str(raw[2].value or '').strip() or None
        obs1  = str(raw[3].value or '').strip()
        obs2  = str(raw[4].value or '').strip() if len(raw)>4 else ''
        obs3  = str(raw[5].value or '').strip() if len(raw)>5 else ''
        price_r = raw[6].value; price_d = dat[6].value
        price = float(price_r) if isinstance(price_r,(int,float)) else (float(price_d) if isinstance(price_d,(int,float)) else 0.0)
        ppl_r = raw[8].value; ppl_d = dat[8].value
        people = int(round(float(ppl_r))) if isinstance(ppl_r,(int,float)) else (int(round(float(ppl_d))) if isinstance(ppl_d,(int,float)) else 1)
        qty    = int(raw[9].value or 1) if raw[9].value else 1
        paid_r = raw[11].value; paid_d = dat[11].value
        paid = float(paid_d) if isinstance(paid_d,(int,float)) else (float(paid_r) if isinstance(paid_r,(int,float)) else 0.0)
        if not title: title = obs1 if obs1 and obs1!='-' else 'Item'
        notes = ' . '.join(p for p in [obs1,obs2,obs3] if p and p not in ('-','')) or None
        if etype == 'Hotel': etype = 'Hospedagem'
        if etype == 'Passeios e Ingressos': etype = 'Passeios'
        if not etype:
            tl = title.lower()
            if any(w in tl for w in ['hotel','pousada','vista do vale']): etype='Hospedagem'
            elif any(w in tl for w in ['voo','aluguel carro','traslado','transfer']): etype='Transporte'
        e = {"id":f"{pfx}-e{eid:02d}","isActive":True,"title":title}
        if etype: e["type"] = etype
        if notes: e["notes"] = notes
        e.update({"price":round(price,2),"taxes":0.0,"people":people,"quantity":qty,
                  "currency":"BRL","exchangeRateToBase":1.0,"paidAmount":round(paid,2)})
        expenses.append(e); eid += 1
    return expenses

expenses = parse_gastos_gramado(wb['Gastos'], wb_data['Gastos'], PFX)

files = sorted([f for f in os.listdir(FOLDER) if not f.startswith('.') and os.path.isfile(os.path.join(FOLDER,f))])
attachments = [{"id":f"{PFX}-att{i+1:02d}","file":f} for i,f in enumerate(files)]

trip = {
    "schemaVersion":1,"id":PFX,"title":"2022-01 Gramado",
    "startDate":"2022-02-27","endDate":"2022-03-04",
    "baseCurrency":"BRL","people":5,"rateDecimalDigits":2,"myMapsUrl":maps_url,
    "itinerarySlotsPerDay":7,
    "itineraryVersions":[{"id":f"{PFX}-v1","name":"Versao 1","bankRows":14,
                          "itinerary":itinerary,"bankActivities":bank_activities}],
    "activeVersionId":f"{PFX}-v1",
    "tasks":[],"links":links,"expenses":expenses,"currencyRates":[],"attachments":attachments
}

out = os.path.join(FOLDER,'trip.json')
with open(out,'w',encoding='utf-8') as f: json.dump(trip,f,ensure_ascii=False,indent=2)
print(f"OK: {out}")
print(f"  {len(itinerary)} dias | {sum(len(d['activities']) for d in itinerary)} ativ | {len(bank_activities)} banco | {len(links)} dicas | {len(expenses)} gastos | {len(attachments)} anexos")
