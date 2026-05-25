import openpyxl, json, os, re, sys
from import_utils import build_merge_map, make_activity_fn
sys.stdout.reconfigure(encoding='utf-8')

FOLDER = r'G:\Meu Drive\Viagens\2020-11 Torres del Paine - CANCELADA'
XLSX   = FOLDER + r'\Viagem Torres del Paine.xlsx'
PFX    = 'torrespaine-2020'

wb      = openpyxl.load_workbook(XLSX)
wb_data = openpyxl.load_workbook(XLSX, data_only=True)
ws      = wb['Roteiro']
merge_map = build_merge_map(ws)
act = make_activity_fn(ws, merge_map, first_col=4)

DATES = ["2020-11-17","2020-11-18","2020-11-19","2020-11-20","2020-11-21","2020-11-22"]
SUMMARIES = ["Ida","Traslado","Base das Torres","Glaciar Grey","Dia Livre","Volta"]

DAYS_DEF = [
    (3, 0, [(9,'a01','Transporte'),(10,'a02','Transporte')]),
    (4, 1, [(7,'a03','Transporte'),(10,'a04','Hospedagem')]),
    (5, 2, [(4,'a05','Passeio'),(10,'a06','Hospedagem')]),
    (6, 3, [(4,'a07','Passeio'),(10,'a08','Hospedagem')]),
    (7, 4, [(4,'a09','Transporte'),(5,'a10','Passeio'),(7,'a11','Passeio'),(10,'a12','Hospedagem')]),
    (8, 5, [(4,'a13','Passeio'),(6,'a14','Transporte'),(7,'a15','Transporte'),(9,'a16','Transporte'),(10,'a17','Transporte')]),
]

itinerary = []
for (row, di, cols) in DAYS_DEF:
    activities = []
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = act(f"{PFX}-{sid}", row, col, atype)
        if a: activities.append(a)
    itinerary.append({"id":f"{PFX}-d{di+1:02d}","date":DATES[di],
                      "title":f"Dia {di+1}","summary":SUMMARIES[di],"activities":activities})

ws_d = wb['Dicas']
links, seen_urls, lid = [], set(), 1
maps_url = ""
for row in ws_d.iter_rows(min_row=2):
    tv = row[0].value; uv = row[1].value if len(row)>1 else None
    if not tv or not uv: continue
    url = re.sub(r'\+[A-Z]\d+$','',str(uv).strip())
    title = str(tv).strip()
    if not url.startswith('http'): continue
    if title == 'Mapa': maps_url = url; continue
    if url in seen_urls: continue
    seen_urls.add(url)
    links.append({"id":f"{PFX}-l{lid:02d}","title":title,"url":url}); lid += 1

def parse_gastos_torres2020(ws_g, ws_g_data, pfx):
    expenses = []; eid = 1
    for rr, rd in zip(ws_g.iter_rows(min_row=2), ws_g_data.iter_rows(min_row=2)):
        raw = list(rr); dat = list(rd)
        if str(raw[0].value or '').strip() not in (chr(9658), chr(9654), '>'): continue
        title = str(raw[1].value or '').strip()
        etype = str(raw[2].value or '').strip() or None
        obs1  = str(raw[3].value or '').strip()
        obs2  = str(raw[4].value or '').strip() if len(raw)>4 else ''
        obs3  = str(raw[5].value or '').strip() if len(raw)>5 else ''
        price_r = raw[6].value; price_d = dat[6].value
        price = float(price_r) if isinstance(price_r,(int,float)) else (float(price_d) if isinstance(price_d,(int,float)) else 0.0)
        taxes_r = raw[7].value
        taxes = float(taxes_r) if isinstance(taxes_r,(int,float)) else 0.0
        ppl_r = raw[8].value; ppl_d = dat[8].value
        people = int(round(float(ppl_r))) if isinstance(ppl_r,(int,float)) else (int(round(float(ppl_d))) if isinstance(ppl_d,(int,float)) else 1)
        qty    = int(raw[9].value or 1) if raw[9].value else 1
        moeda_raw  = str(raw[11].value or '1').strip()
        moeda_calc = dat[11].value
        currency, rate = 'BRL', 1.0
        if 'CLP' in moeda_raw:
            currency = 'CLP'; rate = float(moeda_calc) if isinstance(moeda_calc,(int,float)) else 0.0
        elif 'Dolar' in moeda_raw or chr(243) in moeda_raw:
            currency = 'USD'; rate = float(moeda_calc) if isinstance(moeda_calc,(int,float)) else 0.0
        elif moeda_raw == '1':
            currency = 'BRL'; rate = 1.0
        else:
            try:
                r = float(moeda_raw)
                if r > 2.0: currency, rate = 'USD', r
                elif 0.001 <= r <= 0.1: currency, rate = 'CLP', r
            except: pass
        paid_r = raw[13].value; paid_d = dat[13].value
        paid = float(paid_d) if isinstance(paid_d,(int,float)) else (float(paid_r) if isinstance(paid_r,(int,float)) else 0.0)
        if not title: title = obs1 if obs1 and obs1!='-' else (obs2 if obs2 and obs2!='-' else 'Item')
        notes = ' . '.join(p for p in [obs1,obs2,obs3] if p and p not in ('-','')) or None
        if etype == 'Hotel': etype = 'Hospedagem'
        if etype in ('Atividades','Passeios e Ingressos'): etype = 'Passeios'
        if not etype:
            tl = title.lower()
            if any(w in tl for w in ['lodge','hostel','hotel','pousada','refug']): etype='Hospedagem'
            elif any(w in tl for w in ['voo','aluguel','traslado','transfer']): etype='Transporte'
            elif any(w in tl for w in ['kayak','trekking','ingresso','parque','naveg']): etype='Passeios'
        e = {"id":f"{pfx}-e{eid:02d}","isActive":True,"title":title}
        if etype: e["type"] = etype
        if notes: e["notes"] = notes
        e.update({"price":round(price,2),"taxes":round(taxes,2),"people":people,"quantity":qty,
                  "currency":currency,"exchangeRateToBase":round(rate,6),"paidAmount":round(paid,2)})
        expenses.append(e); eid += 1
    return expenses

expenses = parse_gastos_torres2020(wb['Gastos'], wb_data['Gastos'], PFX)

files = sorted([f for f in os.listdir(FOLDER) if not f.startswith('.') and os.path.isfile(os.path.join(FOLDER,f))])
attachments = [{"id":f"{PFX}-att{i+1:02d}","file":f} for i,f in enumerate(files)]

trip = {
    "schemaVersion":1,"id":PFX,"title":"2020-11 Torres del Paine - CANCELADA",
    "startDate":"2020-11-17","endDate":"2020-11-22",
    "baseCurrency":"BRL","people":2,"rateDecimalDigits":2,"myMapsUrl":maps_url,
    "itinerarySlotsPerDay":7,
    "itineraryVersions":[{"id":f"{PFX}-v1","name":"Vers" + chr(227) + "o 1","bankRows":0,
                          "itinerary":itinerary,"bankActivities":[]}],
    "activeVersionId":f"{PFX}-v1",
    "tasks":[],"links":links,"expenses":expenses,"currencyRates":[],"attachments":attachments
}

out = os.path.join(FOLDER,'trip.json')
with open(out,'w',encoding='utf-8') as f: json.dump(trip,f,ensure_ascii=False,indent=2)
print(f"OK: {out}")
print(f"  {len(itinerary)} dias | {sum(len(d['activities']) for d in itinerary)} ativ | {len(links)} dicas | {len(expenses)} gastos | {len(attachments)} anexos")
