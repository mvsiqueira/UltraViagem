import openpyxl, json, os, re, sys
from import_utils import build_merge_map, make_activity_fn
sys.stdout.reconfigure(encoding='utf-8')

XLSX   = r'G:\Meu Drive\Viagens\2023-04 Chapada Diamantina\Viagem Chapada Diamantina2.xlsx'
FOLDER = r'G:\Meu Drive\Viagens\2023-04 Chapada Diamantina'
PFX    = 'chapada-2023'

wb      = openpyxl.load_workbook(XLSX)
wb_data = openpyxl.load_workbook(XLSX, data_only=True)
ws      = wb['Roteiro']
merge_map = build_merge_map(ws)
act = make_activity_fn(ws, merge_map, first_col=5)

DATES     = ["2023-04-15","2023-04-16","2023-04-17","2023-04-18",
             "2023-04-19","2023-04-20","2023-04-21","2023-04-22"]
SUMMARIES = ["IDA","Mucuge 1","Mucuge 2","Vale do Capao 1",
             "Vale do Capao 2","Lencois 1","Lencois 2","VOLTA"]

DAYS_DEF = [
    (3, 0, [(5,'a01','Transporte'),(7,'a02','Transporte'),(8,'a03','Refeicao'),
            (9,'a04','Transporte'),(12,'a05','Hospedagem')]),
    (4, 1, [(5,'a06','Passeio'),(6,'a07','Passeio'),(7,'a08','Refeicao'),
            (8,'a09','Passeio'),(10,'a10','Passeio'),(12,'a11','Hospedagem')]),
    (5, 2, [(5,'a12','Transporte'),(6,'a13','Trilha'),(10,'a14','Transporte'),
            (12,'a15','Hospedagem')]),
    (6, 3, [(5,'a16','Passeio'),(6,'a17','Passeio'),(7,'a18','Refeicao'),
            (8,'a19','Transporte'),(9,'a20','Passeio'),(12,'a21','Hospedagem')]),
    (7, 4, [(5,'a22','Trilha'),(10,'a23','Passeio'),(12,'a24','Hospedagem')]),
    (8, 5, [(5,'a25','Transporte'),(6,'a26','Passeio'),(7,'a27','Refeicao'),
            (8,'a28','Passeio'),(10,'a29','Passeio'),(12,'a30','Hospedagem')]),
    (9, 6, [(5,'a31','Trilha'),(8,'a32','Refeicao'),(9,'a33','Passeio'),
            (10,'a34','Passeio'),(12,'a35','Hospedagem')]),
    (10,7, [(5,'a36','Transporte'),(7,'a37','Refeicao'),(9,'a38','Transporte'),
            (10,'a39','Transporte')]),
]

itinerary = []
for (row, di, cols) in DAYS_DEF:
    activities = []
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = act(f"{PFX}-{sid}", row, col, atype)
        if a: activities.append(a)
    itinerary.append({"id":f"{PFX}-d{di+1:02d}","date":DATES[di],
                      "title":f"Dia {di+1}","summary":SUMMARIES[di],"activities":activities})

BANK_DEF = [(12, 0, [(5,'b01','Trilha')])]
bank_activities = []
for (row, br, cols) in BANK_DEF:
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = act(f"{PFX}-{sid}", row, col, atype, bankrow=br)
        if a: bank_activities.append(a)

ws_d = wb['Dicas']
links, seen_urls, lid = [], set(), 1
maps_url = ""
for row in ws_d.iter_rows(min_row=2):
    tv = row[0].value; uv = row[1].value if len(row)>1 else None
    if not tv or not uv: continue
    url = re.sub(r'\+[A-Z]\d+$','',str(uv).strip())
    title = str(tv).strip()
    if title == 'Mapa': maps_url = url; continue
    if url in seen_urls: continue
    seen_urls.add(url)
    links.append({"id":f"{PFX}-l{lid:02d}","title":title,"url":url}); lid += 1

def parse_gastos_brl(ws_g, ws_g_data, pfx, pi, ti, ppli, qi, pagi):
    expenses = []; eid = 1
    for rr, rd in zip(ws_g.iter_rows(min_row=2), ws_g_data.iter_rows(min_row=2)):
        raw = list(rr); dat = list(rd)
        if str(raw[0].value or '').strip() not in ('>', chr(9658), chr(9654)): continue
        title   = str(raw[1].value or '').strip()
        etype   = str(raw[2].value or '').strip() or None
        obs1    = str(raw[3].value or '').strip()
        obs2    = str(raw[4].value or '').strip() if len(raw)>4 else ''
        obs3    = str(raw[5].value or '').strip() if len(raw)>5 else ''
        price_r = raw[pi].value; price_d = dat[pi].value
        price = float(price_r) if isinstance(price_r,(int,float)) else (float(price_d) if isinstance(price_d,(int,float)) else 0.0)
        taxes_r = raw[ti].value if ti is not None else None
        taxes = float(taxes_r) if isinstance(taxes_r,(int,float)) else 0.0
        people = int(round(float(raw[ppli].value))) if isinstance(raw[ppli].value,(int,float)) else 1
        qty    = int(raw[qi].value or 1) if raw[qi].value else 1
        paid_d = dat[pagi].value; paid_r = raw[pagi].value
        paid = float(paid_d) if isinstance(paid_d,(int,float)) else (float(paid_r) if isinstance(paid_r,(int,float)) else 0.0)
        if not title: title = obs1 if obs1 and obs1!='-' else (obs2 if obs2 and obs2!='-' else 'Item')
        notes = ' . '.join(p for p in [obs1,obs2,obs3] if p and p not in ('-','')) or None
        if etype == 'Hotel': etype = 'Hospedagem'
        if etype == 'Atividades': etype = 'Passeios'
        if not etype:
            tl = title.lower()
            if any(w in tl for w in ['hotel','pousada','glamping','hostal','hostel']): etype='Hospedagem'
            elif any(w in tl for w in ['voo','voo ','voo','aluguel','traslado','transfer','onibus','Bus']): etype='Transporte'
            elif any(w in tl for w in ['ingresso','excursao','guia','trilha']): etype='Passeios'
        e = {"id":f"{pfx}-e{eid:02d}","isActive":True,"title":title}
        if etype: e["type"] = etype
        if notes: e["notes"] = notes
        e.update({"price":round(price,2),"taxes":round(taxes,2),"people":people,"quantity":qty,
                  "currency":"BRL","exchangeRateToBase":1.0,"paidAmount":round(paid,2)})
        expenses.append(e); eid += 1
    return expenses

expenses = parse_gastos_brl(wb['Gastos'], wb_data['Gastos'], PFX, 6, 7, 8, 9, 11)

files = sorted([f for f in os.listdir(FOLDER) if not f.startswith('.') and os.path.isfile(os.path.join(FOLDER,f))])
attachments = [{"id":f"{PFX}-att{i+1:02d}","file":f} for i,f in enumerate(files)]

trip = {
    "schemaVersion":1,"id":PFX,"title":"2023-04 Chapada Diamantina",
    "startDate":"2023-04-15","endDate":"2023-04-22",
    "baseCurrency":"BRL","people":2,"rateDecimalDigits":2,"myMapsUrl":maps_url,
    "itinerarySlotsPerDay":8,
    "itineraryVersions":[{"id":f"{PFX}-v1","name":"Versao 1","bankRows":1,
                          "itinerary":itinerary,"bankActivities":bank_activities}],
    "activeVersionId":f"{PFX}-v1",
    "tasks":[],"links":links,"expenses":expenses,"currencyRates":[],"attachments":attachments
}

out = os.path.join(FOLDER,'trip.json')
with open(out,'w',encoding='utf-8') as f: json.dump(trip,f,ensure_ascii=False,indent=2)
print(f"OK: {out}")
print(f"  {len(itinerary)} dias | {sum(len(d['activities']) for d in itinerary)} ativ | {len(bank_activities)} banco | {len(links)} dicas | {len(expenses)} gastos | {len(attachments)} anexos")
