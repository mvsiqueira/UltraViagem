import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'G:\Meu Drive\Viagens\2025-07 Pucón\Viagem Pucón.xlsx')

# Full cell text for Roteiro
ws = wb['Roteiro']
print("=== ROTEIRO - valores completos ===")
for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
    for cell in row:
        if cell.value and str(cell.value).strip() not in ('', '-'):
            fg = cell.fill.fgColor
            if fg.type == 'rgb':
                color = fg.rgb
            elif fg.type == 'theme':
                color = f"theme{fg.theme}_tint{fg.tint:.4f}"
            else:
                color = fg.type
            v = str(cell.value).replace('\n', ' | ')
            print(f"  [{cell.row},{cell.column}] {cell.coordinate}: {v!r}  color={color}")

# Dicas
print()
print("=== DICAS ===")
ws_d = wb['Dicas']
for row in ws_d.iter_rows(min_row=2):
    title = row[0].value
    url = row[1].value if len(row) > 1 else None
    if title or url:
        print(f"  titulo={title!r}  url={url!r}")

# Gastos
print()
print("=== GASTOS ===")
ws_g = wb['Gastos']
for i, row in enumerate(ws_g.iter_rows(min_row=1), start=1):
    vals = [c.value for c in row]
    if i == 1:
        print(f"  HEADER: {vals}")
        continue
    marker = str(vals[0] or '').strip()
    if marker in ('►', '▶', '>'):
        print(f"  row{i}: {[str(v or '')[:40] for v in vals]}")
