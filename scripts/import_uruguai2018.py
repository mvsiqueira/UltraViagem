import openpyxl, json, os, re, sys
from import_utils import build_merge_map, make_activity_fn
sys.stdout.reconfigure(encoding='utf-8')

FOLDER = r'G:\Meu Drive\Viagens\2018-11 Uruguai'
XLSX   = FOLDER + r'\Viagem Uruguai.xlsx'
PFX    = 'uruguai-2018'

wb      = openpyxl.load_workbook(XLSX)
wb_data = openpyxl.load_workbook(XLSX, data_only=True)
ws      = wb['Roteiro']
merge_map = build_merge_map(ws)
act = make_activity_fn(ws, merge_map, first_col=4)

DATES     = ["2018-11-02","2018-11-03","2018-11-04","2018-11-05","2018-11-06"]
SUMMARIES = ["Ida","Ciudad Vieja","Pocitos","Col" + chr(244) + "nia do Sacramento","Volta"]

DAYS_DEF = [
    (3, 0, [(7,'a01','Transporte')]),
    (4, 1, [(4,'a02','Passeio'),(5,'a03','Passeio'),(6,'a04','Refeicao'),(7,'a05','Passeio')]),
    (5, 2, [(4,'a06','Passeio'),(6,'a07','Refeicao'),(7,'a08','Passeio'),(8,'a09','Passeio'),(9,'a10','Passeio')]),
    (6, 3, [(4,'a11','Passeio')]),
    (7, 4, [(4,'a12','Transporte')]),
]

itinerary = []
for (row, di, cols) in DAYS_DEF:
    activities = []
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = act(f"{PFX}-{sid}", row, col, atype)
        if a: activities.append(a)
    itinerary.append({"id":f"{PFX}-d{di+1:02d}","date":DATES[di],
                      "title":f"Dia {di+1}","summary":SUMMARIES[di],"activities":activities})

ws_d = wb['Dicas']
links, seen_urls, lid = [], set(), 1
maps_url = ""
for row in ws_d.iter_rows(min_row=2):
    tv = row[0].value
    # Dicas do Uruguai pode ter 2 ou 3 colunas — URL sempre na última coluna não-vazia
    uv = None
    for c in row[1:]:
        if c.value and str(c.value).strip().startswith('http'):
            uv = c.value; break
    if not tv or not uv: continue
    url = re.sub(r'\+[A-Z]\d+$','',str(uv).strip())
    title = str(tv).strip()
    if not url.startswith('http'): continue
    if title == 'Mapa': maps_url = url; continue
    if url in seen_urls: continue
    seen_urls.add(url)
    links.append({"id":f"{PFX}-l{lid:02d}","title":title,"url":url}); lid += 1

def parse_gastos_multi_brl(ws_g, ws_g_data, pfx, pagi_idx):
    """Converte preços multi-moeda (Peso/Dólar/BRL) para R$ usando taxa da col L (data_only)."""
    expenses = []; eid = 1
    for rr, rd in zip(ws_g.iter_rows(min_row=2), ws_g_data.iter_rows(min_row=2)):
        raw = list(rr); dat = list(rd)
        if str(raw[0].value or '').strip() not in ('>', chr(9658), chr(9654)): continue
        title = str(raw[1].value or '').strip()
        etype = str(raw[2].value or '').strip() or None
        obs1  = str(raw[3].value or '').strip()
        obs2  = str(raw[4].value or '').strip() if len(raw)>4 else ''
        obs3  = str(raw[5].value or '').strip() if len(raw)>5 else ''
        price_g  = dat[6].value
        taxes_g  = dat[7].value
        exchange = dat[11].value
        people_v = dat[8].value
        qty_v    = dat[9].value
        paid_v   = dat[pagi_idx].value if pagi_idx < len(dat) else None

        if not isinstance(exchange,(int,float)) or exchange == 0: exchange = 1.0
        price_g = float(price_g) if isinstance(price_g,(int,float)) else 0.0
        taxes_g = float(taxes_g) if isinstance(taxes_g,(int,float)) else 0.0
        price  = round(price_g * exchange, 2)
        taxes  = round(taxes_g * exchange, 2)
        people = int(round(float(people_v))) if isinstance(people_v,(int,float)) else 1
        qty    = int(qty_v) if isinstance(qty_v,(int,float)) else 1
        paid   = float(paid_v) if isinstance(paid_v,(int,float)) else 0.0

        if not title: title = obs1 if obs1 and obs1!='-' else (obs2 if obs2 and obs2!='-' else 'Item')
        notes = ' . '.join(p for p in [obs1,obs2,obs3] if p and p not in ('-','')) or None
        if etype == 'Hotel': etype = 'Hospedagem'
        if etype in ('Atividades','Passeios e Ingressos','Refeição'): etype = 'Passeios' if 'tividade' in (etype or '') else etype
        if etype == 'Refeição': etype = 'Refeicao'
        if not etype:
            tl = title.lower()
            if any(w in tl for w in ['hotel','pousada']): etype='Hospedagem'
            elif any(w in tl for w in ['v' + chr(244) + 'o','traslado','passagem']): etype='Transporte'
            elif any(w in tl for w in ['bus','tour','museu']): etype='Passeios'
            elif any(w in tl for w in ['almo' + chr(231) + 'o','jantar']): etype='Refeicao'
        e = {"id":f"{pfx}-e{eid:02d}","isActive":True,"title":title}
        if etype: e["type"] = etype
        if notes: e["notes"] = notes
        e.update({"price":price,"taxes":taxes,"people":people,"quantity":qty,
                  "currency":"BRL","exchangeRateToBase":1.0,"paidAmount":paid})
        expenses.append(e); eid += 1
    return expenses

# pagi_idx=14 = col O = Pago em R$
expenses = parse_gastos_multi_brl(wb['Gastos'], wb_data['Gastos'], PFX, 14)

files = sorted([f for f in os.listdir(FOLDER) if not f.startswith('.') and os.path.isfile(os.path.join(FOLDER,f))])
attachments = [{"id":f"{PFX}-att{i+1:02d}","file":f} for i,f in enumerate(files)]

trip = {
    "schemaVersion":1,"id":PFX,"title":"2018-11 Uruguai",
    "startDate":"2018-11-02","endDate":"2018-11-06",
    "baseCurrency":"BRL","people":2,"rateDecimalDigits":2,"myMapsUrl":maps_url,
    "itinerarySlotsPerDay":7,
    "itineraryVersions":[{"id":f"{PFX}-v1","name":"Vers" + chr(227) + "o 1","bankRows":3,
                          "itinerary":itinerary,"bankActivities":[]}],
    "activeVersionId":f"{PFX}-v1",
    "tasks":[],"links":links,"expenses":expenses,"currencyRates":[],"attachments":attachments
}

out = os.path.join(FOLDER,'trip.json')
with open(out,'w',encoding='utf-8') as f: json.dump(trip,f,ensure_ascii=False,indent=2)
print(f"OK: {out}")
print(f"  {len(itinerary)} dias | {sum(len(d['activities']) for d in itinerary)} ativ | {len(links)} dicas | {len(expenses)} gastos | {len(attachments)} anexos")
