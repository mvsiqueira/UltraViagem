import openpyxl, json, colorsys, os, re, sys
sys.stdout.reconfigure(encoding='utf-8')

XLSX   = r'G:\Meu Drive\Viagens\2025-03 El Chaltén\Viagem El Chatén.xlsx'
FOLDER = r'G:\Meu Drive\Viagens\2025-03 El Chaltén'

THEME = {0:'000000',1:'FFFFFF',2:'1F497D',3:'EEECE1',
         4:'4F81BD',5:'C0504D',6:'9BBB59',7:'8064A2',
         8:'4BACC6',9:'F79646',10:'0000FF',11:'800080'}

def apply_tint(h6, t):
    r,g,b = int(h6[0:2],16)/255,int(h6[2:4],16)/255,int(h6[4:6],16)/255
    hh,l,s = colorsys.rgb_to_hls(r,g,b)
    l = l*(1-t)+t if t>0 else l*(1+t)
    l = max(0.0,min(1.0,l))
    r2,g2,b2 = colorsys.hls_to_rgb(hh,l,s)
    return f"#{round(r2*255):02X}{round(g2*255):02X}{round(b2*255):02X}"

def cell_color(cell):
    fg = cell.fill.fgColor
    if cell.fill.fill_type != 'solid': return '#000000'
    if fg.type == 'rgb':
        rgb = fg.rgb[-6:].upper()
        if rgb == '000000' and fg.rgb[:2] == '00': return '#000000'  # transparent → black
        return f"#{rgb}"
    if fg.type == 'theme':
        b = THEME.get(fg.theme,'000000')
        return apply_tint(b,fg.tint) if fg.tint != 0 else f"#{b.upper()}"
    return '#000000'

wb      = openpyxl.load_workbook(XLSX)
wb_data = openpyxl.load_workbook(XLSX, data_only=True)
ws      = wb['Roteiro']

merge_map = {}
for mc in ws.merged_cells.ranges:
    for r in range(mc.min_row, mc.max_row+1):
        for c in range(mc.min_col, mc.max_col+1):
            merge_map[(r,c)] = mc

FIRST_COL = 3

def parse_text(value):
    text = str(value or '').strip()
    if not text or text == '-': return None, None
    for sep in ('|','\n'):
        if sep in text:
            parts = [p.strip() for p in text.split(sep)]
            return parts[0], ' · '.join(p for p in parts[1:] if p)
    return text, ''

def make_activity(aid, row, col, atype, bankrow=0):
    mc = merge_map.get((row, col))
    if mc:
        min_c, max_c = mc.min_col, mc.max_col
        real_cell = ws.cell(row=mc.min_row, column=mc.min_col)
    else:
        min_c = max_c = col
        real_cell = ws.cell(row=row, column=col)
    title, details = parse_text(real_cell.value)
    if not title: return None
    color = cell_color(real_cell)
    a = {"id": aid, "title": title, "type": atype, "color": color, "icon": "",
         "startSlot": min_c - FIRST_COL, "durationSlots": max_c - min_c + 1, "bankRow": bankrow}
    if details: a["details"] = details
    return a

# Estrutura:
#  C(3)=slot0  D(4)=1  E(5)=2  F(6)=3  G(7)=4  H(8)=5  I(9)=6  J(10)=7  K(11)=8
# Linha 3=Dia1(IDA) ... linha 10=Dia8(VOLTA)
# Banco: linhas 12-15

DATES = ["2025-03-22","2025-03-23","2025-03-24","2025-03-25",
         "2025-03-26","2025-03-27","2025-03-28","2025-03-29"]
DAY_TITLES    = [f"Dia {i+1}" for i in range(8)]
DAY_SUMMARIES = ["IDA","El Chaltén","Fitz Roy","Cerro Torre",
                 "Livre","Dos Glaciares","El Calafate","VOLTA"]

DAYS_DEF = [
    # (excel_row, day_idx, [(col, id, type)])
    (3, 0, [(3,'a01','Transporte'),(8,'a02','Transporte'),(11,'a03','Hospedagem')]),
    (4, 1, [(4,'a04','Trilha'),(6,'a05','Refeição'),(7,'a06','Passeio'),(11,'a07','Hospedagem')]),
    (5, 2, [(3,'a08','Trilha'),(10,'a09','Hospedagem'),(11,'a10','Hospedagem')]),
    (6, 3, [(3,'a11','Trilha'),(11,'a12','Hospedagem')]),
    (7, 4, [(11,'a13','Hospedagem')]),
    (8, 5, [(3,'a14','Passeio'),(11,'a15','Hospedagem')]),
    (9, 6, [(3,'a16','Passeio'),(5,'a17','Transporte'),(7,'a18','Refeição'),(11,'a19','Hospedagem')]),
    (10,7, [(3,'a20','Transporte')]),
]

itinerary = []
for (row, di, cols) in DAYS_DEF:
    acts = []
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = make_activity(f"chalten-2025-{sid}", row, col, atype)
        if a: acts.append(a)
    itinerary.append({"id": f"chalten-2025-d{di+1:02d}", "date": DATES[di],
                      "title": DAY_TITLES[di], "summary": DAY_SUMMARIES[di], "activities": acts})

BANK_DEF = [
    (12, 0, [(3,'b01','Trilha'),(5,'b02','Trilha'),(8,'b03','Trilha')]),
    (13, 1, [(3,'b04','Trilha')]),
    (14, 2, [(3,'b05','Trilha')]),
    (15, 3, [(3,'b06','Passeio')]),
]

bank_activities = []
for (row, br, cols) in BANK_DEF:
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = make_activity(f"chalten-2025-{sid}", row, col, atype, bankrow=br)
        if a: bank_activities.append(a)

# Dicas
ws_d = wb['Dicas']
links, seen_urls, lid = [], set(), 1
for row in ws_d.iter_rows(min_row=2):
    tv, uv = row[0].value, (row[1].value if len(row)>1 else None)
    if not tv or not uv: continue
    url = re.sub(r'\+[A-Z]\d+$','',str(uv).strip())
    if url in seen_urls: continue
    seen_urls.add(url)
    links.append({"id":f"chalten-2025-l{lid:02d}","title":str(tv).strip(),"url":url})
    lid += 1

# Gastos
ws_g      = wb['Gastos']
ws_g_data = wb_data['Gastos']
expenses  = []
eid       = 1

rows_raw  = list(ws_g.iter_rows(min_row=2))
rows_data = list(ws_g_data.iter_rows(min_row=2))

def gv(cells, col): return list(cells)[col-1].value if col-1 < len(list(cells)) else None

for row_raw, row_dat in zip(rows_raw, rows_data):
    raw = list(row_raw); dat = list(row_dat)
    marker = str(raw[0].value or '').strip()
    if marker not in ('►','▶'): continue

    title_v = raw[1].value
    title   = str(title_v or '').strip()
    etype   = str(raw[2].value or '').strip() or None
    company = str(raw[3].value or '').strip()
    obs1    = str(raw[5].value or '').strip()
    obs2    = str(raw[6].value or '').strip()
    obs3    = str(raw[7].value or '').strip()

    price_raw = raw[8].value
    price_dat = dat[8].value
    if isinstance(price_raw, (int,float)): price = float(price_raw)
    elif isinstance(price_dat, (int,float)): price = float(price_dat)
    else: price = 0.0

    taxes_raw = raw[9].value
    taxes = float(taxes_raw) if isinstance(taxes_raw,(int,float)) else 0.0

    people = int(raw[10].value or 1)
    qty    = int(raw[11].value or 1)

    moeda_raw  = str(raw[13].value or '1').strip()
    moeda_calc = dat[13].value
    if 'Peso' in moeda_raw or 'peso' in moeda_raw:
        currency, rate = 'CLP', (float(moeda_calc) if isinstance(moeda_calc,(int,float)) else 0.0)
    elif 'Dólar' in moeda_raw or 'Dolar' in moeda_raw:
        currency, rate = 'USD', (float(moeda_calc) if isinstance(moeda_calc,(int,float)) else 0.0)
    else:
        currency, rate = 'BRL', 1.0

    paid_raw  = raw[15].value
    paid_calc = dat[15].value
    if isinstance(paid_calc,(int,float)): paid = float(paid_calc)
    elif isinstance(paid_raw,str) and paid_raw.startswith('='): paid = (price+taxes)*people*qty*rate
    else: paid = 0.0

    if not title:
        company_clean = company if company and company != '-' else ''
        title = f"Transfer ({obs1})" if obs1 and obs1 != '-' else (company_clean or 'Item')

    note_parts = [p for p in [obs1,obs2,obs3] if p and p not in ('-','')]
    notes = ' · '.join(note_parts) if note_parts else None
    company = company if company and company != '-' else None

    if not etype:
        tl = title.lower()
        if any(w in tl for w in ['hotel','pousada','glamping','cabin','cabaña']): etype = 'Hospedagem'
        elif any(w in tl for w in ['traslado','transfer','vôo','voo','aluguel']): etype = 'Transporte'
        elif any(w in tl for w in ['ingresso','excursão','excursao']): etype = 'Passeios'
    if etype == 'Hotel': etype = 'Hospedagem'

    e = {"id":f"chalten-2025-e{eid:02d}","isActive":True,"title":title}
    if etype:   e["type"]    = etype
    if company: e["company"] = company
    if notes:   e["notes"]   = notes
    e["price"]              = round(price,2)
    e["taxes"]              = round(taxes,2)
    e["people"]             = people
    e["quantity"]           = qty
    e["currency"]           = currency
    e["exchangeRateToBase"] = round(rate,6)
    e["paidAmount"]         = round(paid,2)
    expenses.append(e)
    eid += 1

files = sorted([f for f in os.listdir(FOLDER)
                if not f.startswith('.') and os.path.isfile(os.path.join(FOLDER,f))])
attachments = [{"id":f"chalten-2025-att{i+1:02d}","file":f} for i,f in enumerate(files)]

trip = {
    "schemaVersion": 1, "id": "chalten-2025",
    "title": "2025-03 El Chaltén",
    "startDate": "2025-03-22", "endDate": "2025-03-29",
    "baseCurrency": "BRL", "people": 2, "rateDecimalDigits": 2,
    "myMapsUrl": "",
    "itinerarySlotsPerDay": 9,
    "itineraryVersions": [{"id":"chalten-2025-v1","name":"Versão 1","bankRows":4,
                            "itinerary":itinerary,"bankActivities":bank_activities}],
    "activeVersionId": "chalten-2025-v1",
    "tasks":[],"links":links,"expenses":expenses,"currencyRates":[],"attachments":attachments
}

out = os.path.join(FOLDER,'trip.json')
with open(out,'w',encoding='utf-8') as f: json.dump(trip,f,ensure_ascii=False,indent=2)

print(f"OK: {out}")
print(f"  {len(itinerary)} dias | {sum(len(d['activities']) for d in itinerary)} atividades")
print(f"  {len(bank_activities)} banco | {len(links)} dicas | {len(expenses)} gastos | {len(attachments)} anexos")
print("\nAtividades:")
for d in itinerary:
    print(f"  {d['date']} {d['summary']}:")
    for a in d['activities']:
        print(f"    [{a['startSlot']}+{a['durationSlots']}] {a['type']:12} {a['title'][:50]}")
print("\nGastos:")
for e in expenses:
    print(f"  {e['id']}: {e['title'][:35]:35} | {e['currency']} p={e['price']} t={e['taxes']} ppl={e['people']} qty={e['quantity']} rate={e['exchangeRateToBase']} paid={e['paidAmount']}")
