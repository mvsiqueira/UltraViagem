import openpyxl, json, colorsys, os, sys
sys.stdout.reconfigure(encoding='utf-8')

XLSX   = r'G:\Meu Drive\Viagens\2025-07 Pucón\Viagem Pucón.xlsx'
FOLDER = r'G:\Meu Drive\Viagens\2025-07 Pucón'

# ── Cores ──────────────────────────────────────────────────────────────────────
THEME = {0:'000000',1:'FFFFFF',2:'1F497D',3:'EEECE1',
         4:'4F81BD',5:'C0504D',6:'9BBB59',7:'8064A2',
         8:'4BACC6',9:'F79646',10:'0000FF',11:'800080'}

def apply_tint(h6, t):
    r,g,b = int(h6[0:2],16)/255,int(h6[2:4],16)/255,int(h6[4:6],16)/255
    hh,l,s = colorsys.rgb_to_hls(r,g,b)
    l = l*(1-t)+t if t>0 else l*(1+t)
    l = max(0.0, min(1.0,l))
    r2,g2,b2 = colorsys.hls_to_rgb(hh,l,s)
    return f"#{round(r2*255):02X}{round(g2*255):02X}{round(b2*255):02X}"

def cell_color(cell):
    fg = cell.fill.fgColor
    if cell.fill.fill_type != 'solid': return '#000000'
    if fg.type == 'rgb': return f"#{fg.rgb[-6:].upper()}"
    if fg.type == 'theme':
        b = THEME.get(fg.theme, '000000')
        return apply_tint(b, fg.tint) if fg.tint != 0 else f"#{b.upper()}"
    return '#000000'

# ── Workbooks ─────────────────────────────────────────────────────────────────
wb      = openpyxl.load_workbook(XLSX)
wb_data = openpyxl.load_workbook(XLSX, data_only=True)
ws      = wb['Roteiro']

# merge map
merge_map = {}
for mc in ws.merged_cells.ranges:
    for r in range(mc.min_row, mc.max_row+1):
        for c in range(mc.min_col, mc.max_col+1):
            merge_map[(r,c)] = mc

FIRST_COL = 3  # coluna C

def parse_text(value):
    """Retorna (title, details) separando pelo primeiro | ou \n."""
    text = str(value or '').strip()
    if not text or text == '-': return None, None
    for sep in ('|', '\n'):
        if sep in text:
            parts = [p.strip() for p in text.split(sep)]
            return parts[0], ' · '.join(p for p in parts[1:] if p)
    return text, ''

def make_activity(aid, row, col, atype, bankrow=0):
    mc = merge_map.get((row, col))
    if mc:
        min_c, max_c = mc.min_col, mc.max_col
        real_cell = ws.cell(row=mc.min_row, column=mc.min_col)
    else:
        min_c = max_c = col
        real_cell = ws.cell(row=row, column=col)
    title, details = parse_text(real_cell.value)
    if not title: return None
    color = cell_color(real_cell)
    a = {"id": aid, "title": title, "type": atype, "color": color, "icon": "",
         "startSlot": min_c - FIRST_COL, "durationSlots": max_c - min_c + 1, "bankRow": bankrow}
    if details:
        a["details"] = details
    return a

# ── Roteiro ────────────────────────────────────────────────────────────────────
#  Colunas C-K = slots 0-8 (9 slots)
#  C(3)=0  D(4)=1  E(5)=2  F(6)=3  G(7)=4  H(8)=5  I(9)=6  J(10)=7  K(11)=8
#
# row: (excel_row, day_index, [(col, id_suffix, type)])
DAYS_DEF = [
    (3, 0, [(8,'a01','Transporte'), (11,'a02','Hospedagem')]),
    (4, 1, [(3,'a03','Passeio'),    (6,'a04','Refeição'),   (7,'a05','Passeio'),
            (8,'a06','Passeio'),    (11,'a07','Hospedagem')]),
    (5, 2, [(3,'a08','Passeio'),    (6,'a09','Refeição'),   (7,'a10','Cultural'),
            (11,'a11','Hospedagem')]),
    (6, 3, [(4,'a12','Transporte'), (6,'a13','Refeição'),   (7,'a14','Transporte'),
            (8,'a15','Passeio'),    (11,'a16','Hospedagem')]),
    (7, 4, [(3,'a17','Passeio'),    (11,'a18','Hospedagem')]),
    (8, 5, [(3,'a19','Passeio'),    (6,'a20','Refeição'),   (11,'a21','Hospedagem')]),
    (9, 6, [(5,'a22','Transporte'), (6,'a23','Transporte'), (9,'a24','Transporte')]),
]

DATES = ["2025-07-25","2025-07-26","2025-07-27","2025-07-28",
         "2025-07-29","2025-07-30","2025-07-31"]
DAY_TITLES    = [f"Dia {i+1}" for i in range(7)]
DAY_SUMMARIES = ["IDA", "Santiago", "Santiago", "Pucón",
                 "Ski Pucón", "Termas Geométricas", "VOLTA"]

itinerary = []
for (row, di, cols) in DAYS_DEF:
    acts = []
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col:
            continue  # skip non-anchor cells of merges
        a = make_activity(f"pucon-2025-{sid}", row, col, atype)
        if a: acts.append(a)
    itinerary.append({
        "id": f"pucon-2025-d{di+1:02d}",
        "date": DATES[di], "title": DAY_TITLES[di],
        "summary": DAY_SUMMARIES[di], "activities": acts
    })

# ── Banco ──────────────────────────────────────────────────────────────────────
BANK_DEF = [
    (11, 0, [(3,'b01','Passeio')]),
    (12, 1, [(3,'b02','Passeio')]),
]

bank_activities = []
for (row, br, cols) in BANK_DEF:
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = make_activity(f"pucon-2025-{sid}", row, col, atype, bankrow=br)
        if a: bank_activities.append(a)

# ── Dicas ──────────────────────────────────────────────────────────────────────
import re
ws_d = wb['Dicas']
links = []
seen_urls = set()
lid = 1
for row in ws_d.iter_rows(min_row=2):
    title_v = row[0].value
    url_v   = row[1].value if len(row) > 1 else None
    if not title_v or not url_v: continue
    url = str(url_v).strip()
    url = re.sub(r'\+[A-Z]\d+$', '', url)  # remove stray cell refs like "+B2"
    if url in seen_urls: continue
    seen_urls.add(url)
    links.append({"id": f"pucon-2025-l{lid:02d}", "title": str(title_v).strip(), "url": url})
    lid += 1

# ── Gastos ─────────────────────────────────────────────────────────────────────
ws_g      = wb['Gastos']
ws_g_data = wb_data['Gastos']
expenses  = []
eid = 1

def gv(row_cells, col):
    """1-based column, returns None if out of range."""
    cells = list(row_cells)
    return cells[col-1].value if col-1 < len(cells) else None

def gv_data(row_cells, col):
    cells = list(row_cells)
    return cells[col-1].value if col-1 < len(cells) else None

rows_raw  = list(ws_g.iter_rows(min_row=2))
rows_data = list(ws_g_data.iter_rows(min_row=2))

for i, (row_raw, row_dat) in enumerate(zip(rows_raw, rows_data)):
    marker = str(gv(row_raw, 1) or '').strip()
    if marker not in ('►', '▶'): continue

    title   = str(gv(row_raw,  2) or '').strip()
    etype   = str(gv(row_raw,  3) or '').strip() or None
    company = str(gv(row_raw,  4) or '').strip()
    obs1    = str(gv(row_raw,  6) or '').strip()
    obs2    = str(gv(row_raw,  7) or '').strip()
    obs3    = str(gv(row_raw,  8) or '').strip()
    price   = float(gv(row_raw,  9) or 0)
    taxes   = float(gv(row_raw, 10) or 0)
    people  = int(gv(row_raw,  11) or 1)
    qty     = int(gv(row_raw,  12) or 1)

    # currency & exchange rate
    moeda_raw  = str(gv(row_raw, 14) or '1').strip()
    moeda_calc = gv_data(row_dat, 14)

    if 'Peso' in moeda_raw or 'peso' in moeda_raw:
        currency = 'CLP'
        rate = float(moeda_calc) if isinstance(moeda_calc, (int, float)) else 0.0
    elif 'Dólar' in moeda_raw or 'Dolar' in moeda_raw:
        currency = 'USD'
        rate = float(moeda_calc) if isinstance(moeda_calc, (int, float)) else 0.0
    else:
        currency = 'BRL'
        rate = 1.0

    # paid
    paid_raw  = gv(row_raw,  16)
    paid_calc = gv_data(row_dat, 16)
    if isinstance(paid_calc, (int, float)):
        paid = float(paid_calc)
    elif isinstance(paid_raw, str) and paid_raw.startswith('='):
        paid = (price + taxes) * people * qty * rate
    else:
        paid = 0.0

    # notes
    note_parts = [p for p in [obs1, obs2, obs3] if p and p != '-']
    notes = ' · '.join(note_parts) if note_parts else None

    company = company if company and company != '-' else None

    # infer type
    if not etype:
        tl = title.lower()
        if any(w in tl for w in ['hotel','pousada']): etype = 'Hospedagem'
        elif any(w in tl for w in ['aluguel','uber','transfer','traslado','vôo','voo']): etype = 'Transporte'
        elif any(w in tl for w in ['ingresso','excursão','excursao','funicular','zoológico','zoo','museu']): etype = 'Passeios'
        elif etype == 'Hotel': etype = 'Hospedagem'
        elif etype == 'Atividades': etype = 'Passeios'
    elif etype == 'Hotel': etype = 'Hospedagem'
    elif etype == 'Atividades': etype = 'Passeios'

    e = {"id": f"pucon-2025-e{eid:02d}", "isActive": True, "title": title}
    if etype:   e["type"]    = etype
    if company: e["company"] = company
    if notes:   e["notes"]   = notes
    e["price"]              = round(price, 2)
    e["taxes"]              = round(taxes, 2)
    e["people"]             = people
    e["quantity"]           = qty
    e["currency"]           = currency
    e["exchangeRateToBase"] = round(rate, 6)
    e["paidAmount"]         = round(paid, 2)
    expenses.append(e)
    eid += 1

# ── Anexos ─────────────────────────────────────────────────────────────────────
files = sorted([f for f in os.listdir(FOLDER)
                if not f.startswith('.') and os.path.isfile(os.path.join(FOLDER, f))])
attachments = [{"id": f"pucon-2025-att{i+1:02d}", "file": f} for i,f in enumerate(files)]

# ── Trip ───────────────────────────────────────────────────────────────────────
trip = {
    "schemaVersion": 1,
    "id": "pucon-2025",
    "title": "2025-07 Pucón",
    "startDate": "2025-07-25",
    "endDate": "2025-07-31",
    "baseCurrency": "BRL",
    "people": 5,
    "rateDecimalDigits": 2,
    "myMapsUrl": "",
    "itinerarySlotsPerDay": 9,
    "itineraryVersions": [{
        "id": "pucon-2025-v1",
        "name": "Versão 1",
        "bankRows": 2,
        "itinerary": itinerary,
        "bankActivities": bank_activities
    }],
    "activeVersionId": "pucon-2025-v1",
    "tasks": [],
    "links": links,
    "expenses": expenses,
    "currencyRates": [],
    "attachments": attachments
}

out = os.path.join(FOLDER, 'trip.json')
with open(out, 'w', encoding='utf-8') as f:
    json.dump(trip, f, ensure_ascii=False, indent=2)

print(f"OK: {out}")
print(f"  {len(itinerary)} dias | {sum(len(d['activities']) for d in itinerary)} atividades")
print(f"  {len(bank_activities)} banco | {len(links)} dicas | {len(expenses)} gastos | {len(attachments)} anexos")
print()
print("Atividades:")
for d in itinerary:
    print(f"  {d['date']} {d['summary']}:")
    for a in d['activities']:
        print(f"    [{a['startSlot']}+{a['durationSlots']}] {a['type']:12} {a['title'][:45]}")
print()
print("Gastos:")
for e in expenses:
    print(f"  {e['id']}: {e['title'][:35]:35} | {e['currency']} price={e['price']} taxes={e['taxes']} "
          f"ppl={e['people']} qty={e['quantity']} rate={e['exchangeRateToBase']} paid={e['paidAmount']}")
print()
print("Dicas:")
for l in links:
    print(f"  {l['id']}: {l['title'][:40]} | {l['url'][:60]}")
