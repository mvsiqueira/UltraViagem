import openpyxl, json, os, re, sys
from import_utils import build_merge_map, make_activity_fn
sys.stdout.reconfigure(encoding='utf-8')

XLSX   = r'G:\Meu Drive\Viagens\2024-11 Emirados Árabes Unidos\Viagem Emirados Arabes.xlsx'
FOLDER = r'G:\Meu Drive\Viagens\2024-11 Emirados Árabes Unidos'
PFX    = 'emirados-2024'

wb      = openpyxl.load_workbook(XLSX)
wb_data = openpyxl.load_workbook(XLSX, data_only=True)
ws      = wb['Roteiro']
merge_map = build_merge_map(ws)
act = make_activity_fn(ws, merge_map, first_col=4)

DATES = ["2024-11-16","2024-11-17","2024-11-18","2024-11-19","2024-11-20",
         "2024-11-21","2024-11-22","2024-11-23","2024-11-24"]
SUMMARIES = ["Ida","Al Fahidi","Burj Khalifa","Deserto e Palm Jumeirah",
             "Downtown Abu Dhabi","Mesquita e Universidade",
             "Abu Dhabi University","Ilha de Saadiyat","Volta"]

DAYS_DEF = [
    (3, 0, [(4,'a01','Transporte'),(13,'a02','Hospedagem')]),
    (4, 1, [(4,'a03','Passeio'),(7,'a04','Refeicao'),(8,'a05','Passeio'),
            (12,'a06','Passeio'),(13,'a07','Hospedagem')]),
    (5, 2, [(4,'a08','Passeio'),(13,'a09','Hospedagem')]),
    (6, 3, [(4,'a10','Passeio'),(7,'a11','Refeicao'),(8,'a12','Passeio'),
            (13,'a13','Hospedagem')]),
    (7, 4, [(4,'a14','Transporte'),(6,'a15','Passeio'),(13,'a16','Hospedagem')]),
    (8, 5, [(4,'a17','Passeio'),(7,'a18','Refeicao'),(8,'a19','Passeio'),
            (12,'a20','Passeio'),(13,'a21','Hospedagem')]),
    (9, 6, [(4,'a22','Passeio'),(7,'a23','Refeicao'),(8,'a24','Passeio'),
            (10,'a25','Passeio'),(13,'a26','Hospedagem')]),
    (10,7, [(4,'a27','Passeio'),(7,'a28','Refeicao'),(8,'a29','Passeio'),
            (10,'a30','Passeio'),(13,'a31','Hospedagem')]),
    (11,8, [(4,'a32','Transporte'),(6,'a33','Transporte')]),
]

itinerary = []
for (row, di, cols) in DAYS_DEF:
    activities = []
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = act(f"{PFX}-{sid}", row, col, atype)
        if a: activities.append(a)
    itinerary.append({"id":f"{PFX}-d{di+1:02d}","date":DATES[di],
                      "title":f"Dia {di+1}","summary":SUMMARIES[di],"activities":activities})

BANK_DEF = [
    (14, 0, [(4,'b01','Passeio'),(6,'b02','Passeio')]),
]
bank_activities = []
for (row, br, cols) in BANK_DEF:
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = act(f"{PFX}-{sid}", row, col, atype, bankrow=br)
        if a: bank_activities.append(a)

ws_d = wb['Dicas']
links, seen_urls, lid = [], set(), 1
maps_url = ""
for row in ws_d.iter_rows(min_row=2):
    tv = row[0].value; uv = row[1].value if len(row)>1 else None
    if not tv or not uv: continue
    url = re.sub(r'\+[A-Z]\d+$','',str(uv).strip())
    title = str(tv).strip()
    if title == 'Mapa': maps_url = url; continue
    if url in seen_urls: continue
    seen_urls.add(url)
    links.append({"id":f"{PFX}-l{lid:02d}","title":title,"url":url}); lid += 1

# Gastos Emirados: price@7, taxes@8, pessoas@9, qty@10, moeda@12, pago@14
def parse_gastos_emirados(ws_g, ws_g_data, pfx):
    expenses = []; eid = 1
    for rr, rd in zip(ws_g.iter_rows(min_row=2), ws_g_data.iter_rows(min_row=2)):
        raw = list(rr); dat = list(rd)
        if str(raw[0].value or '').strip() not in (chr(9658), chr(9654)): continue
        title  = str(raw[1].value or '').strip()
        etype  = str(raw[2].value or '').strip() or None
        obs1   = str(raw[3].value or '').strip()
        obs2   = str(raw[4].value or '').strip() if len(raw)>4 else ''
        obs3   = str(raw[5].value or '').strip() if len(raw)>5 else ''
        price_r = raw[7].value; price_d = dat[7].value
        price = float(price_r) if isinstance(price_r,(int,float)) else (float(price_d) if isinstance(price_d,(int,float)) else 0.0)
        taxes_r = raw[8].value
        taxes = float(taxes_r) if isinstance(taxes_r,(int,float)) else 0.0
        people = int(round(float(raw[9].value))) if isinstance(raw[9].value,(int,float)) else 1
        qty    = int(raw[10].value or 1) if raw[10].value else 1
        moeda_raw  = str(raw[12].value or '1').strip()
        moeda_calc = dat[12].value
        currency, rate = 'BRL', 1.0
        if 'AED' in moeda_raw:
            currency = 'AED'; rate = float(moeda_calc) if isinstance(moeda_calc,(int,float)) else 0.0
        elif 'USD' in moeda_raw:
            currency = 'USD'; rate = float(moeda_calc) if isinstance(moeda_calc,(int,float)) else 0.0
        elif moeda_raw == '1':
            currency = 'BRL'; rate = 1.0
        paid_d = dat[14].value; paid_r = raw[14].value
        paid = float(paid_d) if isinstance(paid_d,(int,float)) else (float(paid_r) if isinstance(paid_r,(int,float)) else 0.0)
        if not title: title = obs1 if obs1 and obs1!='-' else (obs2 if obs2 and obs2!='-' else 'Item')
        notes = ' . '.join(p for p in [obs1,obs2,obs3] if p and p not in ('-','')) or None
        if etype == 'Hotel': etype = 'Hospedagem'
        if etype == 'Atividades': etype = 'Passeios'
        if not etype:
            tl = title.lower()
            if any(w in tl for w in ['hotel','atana','quinta','la quinta']): etype='Hospedagem'
            elif any(w in tl for w in ['voo','transfer','nol']): etype='Transporte'
            elif any(w in tl for w in ['tour','safari','burj','ingresso','card','louvre','kai beach']): etype='Passeios'
        e = {"id":f"{pfx}-e{eid:02d}","isActive":True,"title":title}
        if etype: e["type"] = etype
        if notes: e["notes"] = notes
        e.update({"price":round(price,2),"taxes":round(taxes,2),"people":people,"quantity":qty,
                  "currency":currency,"exchangeRateToBase":round(rate,6),"paidAmount":round(paid,2)})
        expenses.append(e); eid += 1
    return expenses

expenses = parse_gastos_emirados(wb['Gastos'], wb_data['Gastos'], PFX)

files = sorted([f for f in os.listdir(FOLDER) if not f.startswith('.') and os.path.isfile(os.path.join(FOLDER,f))])
attachments = [{"id":f"{PFX}-att{i+1:02d}","file":f} for i,f in enumerate(files)]

trip = {
    "schemaVersion":1,"id":PFX,"title":"2024-11 Emirados Arabes Unidos",
    "startDate":"2024-11-16","endDate":"2024-11-24",
    "baseCurrency":"BRL","people":2,"rateDecimalDigits":2,"myMapsUrl":maps_url,
    "itinerarySlotsPerDay":10,
    "itineraryVersions":[{"id":f"{PFX}-v1","name":"Versao 1","bankRows":1,
                          "itinerary":itinerary,"bankActivities":bank_activities}],
    "activeVersionId":f"{PFX}-v1",
    "tasks":[],"links":links,"expenses":expenses,"currencyRates":[],"attachments":attachments
}

out = os.path.join(FOLDER,'trip.json')
with open(out,'w',encoding='utf-8') as f: json.dump(trip,f,ensure_ascii=False,indent=2)
print(f"OK: {out}")
print(f"  {len(itinerary)} dias | {sum(len(d['activities']) for d in itinerary)} ativ | {len(bank_activities)} banco | {len(links)} dicas | {len(expenses)} gastos | {len(attachments)} anexos")
