import openpyxl, json, os, re, sys
from import_utils import build_merge_map, make_activity_fn
sys.stdout.reconfigure(encoding='utf-8')

FOLDER = r'G:\Meu Drive\Viagens\2019-01 Orlando'
XLSX   = FOLDER + r'\Viagem Orlando.xlsx'
PFX    = 'orlando-2019'

wb      = openpyxl.load_workbook(XLSX)
wb_data = openpyxl.load_workbook(XLSX, data_only=True)
ws      = wb['Roteiro']
merge_map = build_merge_map(ws)
act = make_activity_fn(ws, merge_map, first_col=4)

DATES = [
    "2019-01-25","2019-01-26","2019-01-27","2019-01-28","2019-01-29","2019-01-30","2019-01-31",
    "2019-02-01","2019-02-02","2019-02-03","2019-02-04","2019-02-05","2019-02-06",
]
SUMMARIES = [
    "Ida","DIA 1","DIA 2","DIA 3","DIA 4","DIA 5","DIA 6",
    "DIA 7","DIA 8","DIA 9","DIA 10","DIA 11","Volta",
]

DAYS_DEF = [
    (3,  0, [(6,'a01','Transporte'),(10,'a02','Hospedagem')]),
    (4,  1, [(4,'a03','Passeio'),(7,'a04','Passeio'),(10,'a05','Hospedagem')]),
    (5,  2, [(4,'a06','Passeio'),(10,'a07','Hospedagem')]),
    (6,  3, [(4,'a08','Passeio'),(10,'a09','Hospedagem')]),
    (7,  4, [(4,'a10','Passeio'),(9,'a11','Passeio'),(10,'a12','Hospedagem')]),
    (8,  5, [(4,'a13','Passeio'),(10,'a14','Hospedagem')]),
    (9,  6, [(4,'a15','Passeio'),(9,'a16','Refeicao'),(10,'a17','Hospedagem')]),
    (10, 7, [(4,'a18','Passeio'),(10,'a19','Hospedagem')]),
    (11, 8, [(4,'a20','Passeio'),(9,'a21','Refeicao'),(10,'a22','Hospedagem')]),
    (12, 9, [(4,'a23','Passeio'),(10,'a24','Hospedagem')]),
    (13,10, [(4,'a25','Passeio'),(10,'a26','Hospedagem')]),
    (14,11, [(4,'a27','Passeio'),(8,'a28','Transporte'),(10,'a29','Hospedagem')]),
    (15,12, [(4,'a30','Transporte'),(5,'a31','Transporte'),(10,'a32','Hospedagem')]),
]

itinerary = []
for (row, di, cols) in DAYS_DEF:
    activities = []
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = act(f"{PFX}-{sid}", row, col, atype)
        if a: activities.append(a)
    itinerary.append({"id":f"{PFX}-d{di+1:02d}","date":DATES[di],
                      "title":f"Dia {di+1}","summary":SUMMARIES[di],"activities":activities})

ws_d = wb['Dicas']
links, seen_urls, lid = [], set(), 1
maps_url = ""
for row in ws_d.iter_rows(min_row=2):
    tv = row[0].value; uv = row[1].value if len(row)>1 else None
    if not tv or not uv: continue
    url = re.sub(r'\+[A-Z]\d+$','',str(uv).strip())
    title = str(tv).strip()
    if not url.startswith('http'): continue
    if title == 'Mapa': maps_url = url; continue
    if url in seen_urls: continue
    seen_urls.add(url)
    links.append({"id":f"{PFX}-l{lid:02d}","title":title,"url":url}); lid += 1

def parse_gastos_multi_brl(ws_g, ws_g_data, pfx, pagi_idx):
    """Converte preços USD/BRL para R$ usando taxa da col M (data_only).
    Orlando tem 4 colunas de obs (D-G) antes do preço (col H=índice 7)."""
    expenses = []; eid = 1
    for rr, rd in zip(ws_g.iter_rows(min_row=2), ws_g_data.iter_rows(min_row=2)):
        raw = list(rr); dat = list(rd)
        if str(raw[0].value or '').strip() not in ('>', chr(9658), chr(9654)): continue
        title = str(raw[1].value or '').strip()
        etype = str(raw[2].value or '').strip() or None
        # obs: cols D-G = índices 3,4,5,6
        obs_parts = [str(raw[i].value or '').strip() for i in [3,4,5,6] if len(raw)>i and raw[i].value and str(raw[i].value).strip() not in ('-','')]
        # preço na col H = índice 7; câmbio na col M = índice 12
        price_g  = dat[7].value   # col H = Preço
        taxes_g  = dat[8].value   # col I = Taxas
        exchange = dat[12].value  # col M = Moeda (taxa calculada)
        people_v = dat[9].value   # col J
        qty_v    = dat[10].value  # col K
        paid_v   = dat[pagi_idx].value if pagi_idx < len(dat) else None

        if not isinstance(exchange,(int,float)) or exchange == 0: exchange = 1.0
        price_g = float(price_g) if isinstance(price_g,(int,float)) else 0.0
        taxes_g = float(taxes_g) if isinstance(taxes_g,(int,float)) else 0.0
        price  = round(price_g * exchange, 2)
        taxes  = round(taxes_g * exchange, 2)
        people = int(round(float(people_v))) if isinstance(people_v,(int,float)) else 1
        qty    = int(qty_v) if isinstance(qty_v,(int,float)) else 1
        paid   = float(paid_v) if isinstance(paid_v,(int,float)) else 0.0

        if not title:
            title = obs_parts[0] if obs_parts else 'Item'
        notes = ' . '.join(obs_parts) or None
        if etype == 'Hotel': etype = 'Hospedagem'
        if etype == 'Tickets': etype = 'Passeios'
        if not etype:
            tl = title.lower()
            if any(w in tl for w in ['hotel','resort']): etype='Hospedagem'
            elif any(w in tl for w in ['v' + chr(244) + 'o','aluguel','estacionamento']): etype='Transporte'
            elif any(w in tl for w in ['disney','universal','legoland','ticket','pass']): etype='Passeios'
        e = {"id":f"{pfx}-e{eid:02d}","isActive":True,"title":title}
        if etype: e["type"] = etype
        if notes: e["notes"] = notes
        e.update({"price":price,"taxes":taxes,"people":people,"quantity":qty,
                  "currency":"BRL","exchangeRateToBase":1.0,"paidAmount":paid})
        expenses.append(e); eid += 1
    return expenses

# pagi_idx=13 = col N = Subtotal R$ (não há col Pago separada em Orlando)
expenses = parse_gastos_multi_brl(wb['Gastos'], wb_data['Gastos'], PFX, 13)

files = sorted([f for f in os.listdir(FOLDER) if not f.startswith('.') and os.path.isfile(os.path.join(FOLDER,f))])
attachments = [{"id":f"{PFX}-att{i+1:02d}","file":f} for i,f in enumerate(files)]

trip = {
    "schemaVersion":1,"id":PFX,"title":"2019-01 Orlando",
    "startDate":"2019-01-25","endDate":"2019-02-06",
    "baseCurrency":"BRL","people":5,"rateDecimalDigits":2,"myMapsUrl":maps_url,
    "itinerarySlotsPerDay":7,
    "itineraryVersions":[{"id":f"{PFX}-v1","name":"Vers" + chr(227) + "o 1","bankRows":3,
                          "itinerary":itinerary,"bankActivities":[]}],
    "activeVersionId":f"{PFX}-v1",
    "tasks":[],"links":links,"expenses":expenses,"currencyRates":[],"attachments":attachments
}

out = os.path.join(FOLDER,'trip.json')
with open(out,'w',encoding='utf-8') as f: json.dump(trip,f,ensure_ascii=False,indent=2)
print(f"OK: {out}")
print(f"  {len(itinerary)} dias | {sum(len(d['activities']) for d in itinerary)} ativ | {len(links)} dicas | {len(expenses)} gastos | {len(attachments)} anexos")
