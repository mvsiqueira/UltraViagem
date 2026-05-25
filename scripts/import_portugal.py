import openpyxl, json, os, re, sys
from import_utils import build_merge_map, make_activity_fn, parse_expenses
sys.stdout.reconfigure(encoding='utf-8')

XLSX   = r'G:\Meu Drive\Viagens\2023-11 Portugal\Viagem Portugal.xlsx'
FOLDER = r'G:\Meu Drive\Viagens\2023-11 Portugal'
PFX    = 'portugal-2023'

wb      = openpyxl.load_workbook(XLSX)
wb_data = openpyxl.load_workbook(XLSX, data_only=True)
ws      = wb['Roteiro']
merge_map = build_merge_map(ws)
act = make_activity_fn(ws, merge_map, first_col=4)

DATES = ["2023-11-22","2023-11-23","2023-11-24","2023-11-25","2023-11-26",
         "2023-11-27","2023-11-28","2023-11-29","2023-11-30","2023-12-01",
         "2023-12-02","2023-12-03","2023-12-04"]
SUMMARIES = ["Ida","Chegada","Sintra","Lisboa 1","Lisboa 2",
             "Evora","Fatima","Coimbra","Aveiro","Porto 1",
             "Porto 2","Obidos","Volta"]

DAYS_DEF = [
    (3,  0, [(11,'a01','Transporte'),(12,'a02','Hospedagem')]),
    (4,  1, [(4,'a03','Transporte'),(8,'a04','Passeio'),(12,'a05','Hospedagem')]),
    (5,  2, [(4,'a06','Transporte'),(5,'a07','Passeio'),(11,'a08','Transporte'),(12,'a09','Hospedagem')]),
    (6,  3, [(4,'a10','Passeio'),(12,'a11','Hospedagem')]),
    (7,  4, [(4,'a12','Passeio'),(12,'a13','Hospedagem')]),
    (8,  5, [(4,'a14','Transporte'),(6,'a15','Passeio'),(10,'a16','Transporte'),(12,'a17','Hospedagem')]),
    (9,  6, [(4,'a18','Transporte'),(6,'a19','Passeio'),(9,'a20','Transporte'),
             (10,'a21','Passeio'),(11,'a22','Transporte'),(12,'a23','Hospedagem')]),
    (10, 7, [(4,'a24','Passeio'),(12,'a25','Hospedagem')]),
    (11, 8, [(4,'a26','Transporte'),(5,'a27','Passeio'),(11,'a28','Transporte'),(12,'a29','Hospedagem')]),
    (12, 9, [(4,'a30','Passeio'),(12,'a31','Hospedagem')]),
    (13,10, [(4,'a32','Passeio'),(12,'a33','Hospedagem')]),
    (14,11, [(4,'a34','Transporte'),(5,'a35','Passeio'),(6,'a36','Passeio'),(7,'a37','Passeio'),
             (8,'a38','Passeio'),(10,'a39','Transporte'),(11,'a40','Transporte'),(12,'a41','Hospedagem')]),
    (15,12, [(4,'a42','Transporte')]),
]

itinerary = []
for (row, di, cols) in DAYS_DEF:
    activities = []
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = act(f"{PFX}-{sid}", row, col, atype)
        if a: activities.append(a)
    itinerary.append({"id":f"{PFX}-d{di+1:02d}","date":DATES[di],
                      "title":f"Dia {di+1}","summary":SUMMARIES[di],"activities":activities})

BANK_DEF = [
    (17, 0, [(4,'b01','Passeio'),(5,'b02','Passeio'),(6,'b03','Passeio'),(7,'b04','Passeio')]),
    (18, 1, [(4,'b05','Passeio')]),
    (19, 2, [(4,'b06','Passeio')]),
]
bank_activities = []
for (row, br, cols) in BANK_DEF:
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = act(f"{PFX}-{sid}", row, col, atype, bankrow=br)
        if a: bank_activities.append(a)

ws_d = wb['Dicas']
links, seen_urls, lid = [], set(), 1
maps_url = ""
for row in ws_d.iter_rows(min_row=2):
    tv = row[0].value; uv = row[1].value if len(row)>1 else None
    if not tv or not uv: continue
    url = re.sub(r'\+[A-Z]\d+$','',str(uv).strip())
    title = str(tv).strip()
    if title == 'Mapa': maps_url = url; continue
    if url in seen_urls: continue
    seen_urls.add(url)
    links.append({"id":f"{PFX}-l{lid:02d}","title":title,"url":url}); lid += 1

# Portugal: standard col layout, moeda Peso/Euro -> EUR
expenses = parse_expenses(wb['Gastos'], wb_data['Gastos'], PFX,
                          moeda_keys={'Peso':'EUR','Euro':'EUR'})

files = sorted([f for f in os.listdir(FOLDER) if not f.startswith('.') and os.path.isfile(os.path.join(FOLDER,f))])
attachments = [{"id":f"{PFX}-att{i+1:02d}","file":f} for i,f in enumerate(files)]

trip = {
    "schemaVersion":1,"id":PFX,"title":"2023-11 Portugal",
    "startDate":"2023-11-22","endDate":"2023-12-04",
    "baseCurrency":"BRL","people":2,"rateDecimalDigits":2,"myMapsUrl":maps_url,
    "itinerarySlotsPerDay":9,
    "itineraryVersions":[{"id":f"{PFX}-v1","name":"Versao 1","bankRows":3,
                          "itinerary":itinerary,"bankActivities":bank_activities}],
    "activeVersionId":f"{PFX}-v1",
    "tasks":[],"links":links,"expenses":expenses,"currencyRates":[],"attachments":attachments
}

out = os.path.join(FOLDER,'trip.json')
with open(out,'w',encoding='utf-8') as f: json.dump(trip,f,ensure_ascii=False,indent=2)
print(f"OK: {out}")
print(f"  {len(itinerary)} dias | {sum(len(d['activities']) for d in itinerary)} ativ | {len(bank_activities)} banco | {len(links)} dicas | {len(expenses)} gastos | {len(attachments)} anexos")
