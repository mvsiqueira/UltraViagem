import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

import sys
XLSX = sys.argv[1]

wb = openpyxl.load_workbook(XLSX)
print(f"Abas: {wb.sheetnames}")

ws = wb['Roteiro']
print(f"Dimensoes: {ws.dimensions}  max_row={ws.max_row} max_col={ws.max_column}")

print("\n=== MERGED CELLS ===")
for mc in sorted(ws.merged_cells.ranges, key=lambda m: (m.min_row, m.min_col)):
    real = ws.cell(row=mc.min_row, column=mc.min_col)
    v = str(real.value or '').replace('\n', ' | ')[:60]
    fg = real.fill.fgColor
    if fg.type == 'rgb': color = fg.rgb
    elif fg.type == 'theme': color = f"theme{fg.theme}_t{fg.tint:.3f}"
    else: color = fg.type
    print(f"  {str(mc):10} [{mc.min_row},{mc.min_col}:{mc.max_row},{mc.max_col}]  color={color}  '{v}'")

print("\n=== CELULAS PREENCHIDAS (todas as linhas) ===")
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for cell in row:
        if cell.value is None: continue
        fg = cell.fill.fgColor
        if fg.type == 'rgb': color = fg.rgb
        elif fg.type == 'theme': color = f"theme{fg.theme}_t{fg.tint:.3f}"
        else: color = fg.type
        v = str(cell.value).replace('\n', ' | ')
        print(f"  [{cell.row},{cell.column}] {cell.coordinate}: {v!r}  color={color}")

print("\n=== DICAS ===")
ws_d = wb['Dicas']
for row in ws_d.iter_rows(min_row=1):
    vals = [str(c.value or '').strip() for c in row]
    if any(v for v in vals): print(" ", vals)

print("\n=== GASTOS cabecalho + linhas marcadas ===")
ws_g = wb['Gastos']
for i, row in enumerate(ws_g.iter_rows(min_row=1), 1):
    vals = [c.value for c in row]
    marker = str(vals[0] or '').strip()
    if i == 1 or marker in ('►','▶','>'):
        safe = [str(v or '')[:40] for v in vals]
        print(f"  row{i}: {safe}")
