import openpyxl, json, os, re, sys
from import_utils import build_merge_map, make_activity_fn
sys.stdout.reconfigure(encoding='utf-8')

XLSX   = r'G:\Meu Drive\Viagens\2024-07 Porto Seguro\Viagem Porto Seguro.xlsx'
FOLDER = r'G:\Meu Drive\Viagens\2024-07 Porto Seguro'
PFX    = 'portoseguro-2024'

wb      = openpyxl.load_workbook(XLSX)
wb_data = openpyxl.load_workbook(XLSX, data_only=True)
ws      = wb['Roteiro']
merge_map = build_merge_map(ws)
act = make_activity_fn(ws, merge_map, first_col=5)

DATES     = ["2024-07-19","2024-07-20","2024-07-21","2024-07-22",
             "2024-07-23","2024-07-24","2024-07-25","2024-07-26"]
SUMMARIES = ["IDA","Porto Seguro","Trancoso","Caraiva",
             "Eco Resort","Eco Parque","Eco Resort 2","VOLTA"]

DAYS_DEF = [
    (3, 0, [(5,'a01','Transporte'),(7,'a02','Refeicao'),(8,'a03','Passeio'),
            (9,'a04','Passeio'),(11,'a05','Passeio'),(12,'a06','Hospedagem')]),
    (4, 1, [(5,'a07','Passeio'),(7,'a08','Refeicao'),(8,'a09','Passeio'),
            (10,'a10','Transporte'),(12,'a11','Hospedagem')]),
    (5, 2, [(5,'a12','Passeio'),(7,'a13','Refeicao'),(8,'a14','Passeio'),
            (11,'a15','Passeio'),(12,'a16','Hospedagem')]),
    (6, 3, [(5,'a17','Passeio'),(8,'a18','Refeicao'),(9,'a19','Passeio'),
            (11,'a20','Passeio'),(12,'a21','Hospedagem')]),
    (7, 4, [(5,'a22','Transporte'),(7,'a23','Refeicao'),(8,'a24','Passeio'),
            (9,'a25','Hospedagem'),(12,'a26','Hospedagem')]),
    (8, 5, [(5,'a27','Passeio'),(10,'a28','Passeio'),(12,'a29','Hospedagem')]),
    (9, 6, [(5,'a30','Passeio'),(10,'a31','Passeio'),(12,'a32','Hospedagem')]),
    (10,7, [(5,'a33','Transporte'),(6,'a34','Transporte')]),
]

itinerary = []
for (row, di, cols) in DAYS_DEF:
    activities = []
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = act(f"{PFX}-{sid}", row, col, atype)
        if a: activities.append(a)
    itinerary.append({"id":f"{PFX}-d{di+1:02d}","date":DATES[di],
                      "title":f"Dia {di+1}","summary":SUMMARIES[di],"activities":activities})

BANK_DEF = [
    (12, 0, [(5,'b01','Passeio'),(6,'b02','Passeio'),(7,'b03','Passeio'),
             (8,'b04','Passeio'),(9,'b05','Passeio'),(10,'b06','Passeio')]),
    (13, 1, [(5,'b07','Passeio'),(6,'b08','Passeio')]),
    (14, 2, [(5,'b09','Passeio'),(6,'b10','Passeio'),(7,'b11','Passeio'),
             (8,'b12','Passeio'),(9,'b13','Passeio')]),
    (15, 3, [(5,'b14','Passeio'),(6,'b15','Passeio'),(9,'b16','Passeio')]),
]
bank_activities = []
for (row, br, cols) in BANK_DEF:
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = act(f"{PFX}-{sid}", row, col, atype, bankrow=br)
        if a: bank_activities.append(a)

ws_d = wb['Dicas']
links, seen_urls, lid = [], set(), 1
maps_url = ""
for row in ws_d.iter_rows(min_row=2):
    tv = row[0].value; uv = row[1].value if len(row)>1 else None
    if not tv or not uv: continue
    url = re.sub(r'\+[A-Z]\d+$','',str(uv).strip())
    title = str(tv).strip()
    if title == 'Mapa': maps_url = url; continue
    if url in seen_urls: continue
    seen_urls.add(url)
    links.append({"id":f"{PFX}-l{lid:02d}","title":title,"url":url}); lid += 1

def parse_gastos_brl(ws_g, ws_g_data, pfx, pi, ti, ppli, qi, pagi):
    expenses = []; eid = 1
    for rr, rd in zip(ws_g.iter_rows(min_row=2), ws_g_data.iter_rows(min_row=2)):
        raw = list(rr); dat = list(rd)
        if str(raw[0].value or '').strip() not in (chr(9658), chr(9654)): continue
        title  = str(raw[1].value or '').strip()
        etype  = str(raw[2].value or '').strip() or None
        obs1   = str(raw[3].value or '').strip()
        obs2   = str(raw[4].value or '').strip() if len(raw)>4 else ''
        obs3   = str(raw[5].value or '').strip() if len(raw)>5 else ''
        price_r = raw[pi].value; price_d = dat[pi].value
        price = float(price_r) if isinstance(price_r,(int,float)) else (float(price_d) if isinstance(price_d,(int,float)) else 0.0)
        taxes_r = raw[ti].value if ti is not None else None
        taxes = float(taxes_r) if isinstance(taxes_r,(int,float)) else 0.0
        people = int(round(float(raw[ppli].value))) if isinstance(raw[ppli].value,(int,float)) else 1
        qty    = int(raw[qi].value or 1) if raw[qi].value else 1
        paid_d = dat[pagi].value; paid_r = raw[pagi].value
        paid = float(paid_d) if isinstance(paid_d,(int,float)) else (float(paid_r) if isinstance(paid_r,(int,float)) else 0.0)
        if not title: title = obs1 if obs1 and obs1!='-' else (obs2 if obs2 and obs2!='-' else 'Item')
        notes = ' . '.join(p for p in [obs1,obs2,obs3] if p and p not in ('-','')) or None
        if etype == 'Hotel': etype = 'Hospedagem'
        if etype == 'Atividades': etype = 'Passeios'
        if not etype:
            tl = title.lower()
            if any(w in tl for w in ['hotel','pousada','resort','eco resort']): etype='Hospedagem'
            elif any(w in tl for w in ['voo','aluguel','traslado','transfer']): etype='Transporte'
        e = {"id":f"{pfx}-e{eid:02d}","isActive":True,"title":title}
        if etype: e["type"] = etype
        if notes: e["notes"] = notes
        e.update({"price":round(price,2),"taxes":round(taxes,2),"people":people,"quantity":qty,
                  "currency":"BRL","exchangeRateToBase":1.0,"paidAmount":round(paid,2)})
        expenses.append(e); eid += 1
    return expenses

expenses = parse_gastos_brl(wb['Gastos'], wb_data['Gastos'], PFX, 7, 8, 9, 10, 12)

files = sorted([f for f in os.listdir(FOLDER) if not f.startswith('.') and os.path.isfile(os.path.join(FOLDER,f))])
attachments = [{"id":f"{PFX}-att{i+1:02d}","file":f} for i,f in enumerate(files)]

trip = {
    "schemaVersion":1,"id":PFX,"title":"2024-07 Porto Seguro",
    "startDate":"2024-07-19","endDate":"2024-07-26",
    "baseCurrency":"BRL","people":4,"rateDecimalDigits":2,"myMapsUrl":maps_url,
    "itinerarySlotsPerDay":8,
    "itineraryVersions":[{"id":f"{PFX}-v1","name":"Versao 1","bankRows":4,
                          "itinerary":itinerary,"bankActivities":bank_activities}],
    "activeVersionId":f"{PFX}-v1",
    "tasks":[],"links":links,"expenses":expenses,"currencyRates":[],"attachments":attachments
}

out = os.path.join(FOLDER,'trip.json')
with open(out,'w',encoding='utf-8') as f: json.dump(trip,f,ensure_ascii=False,indent=2)
print(f"OK: {out}")
print(f"  {len(itinerary)} dias | {sum(len(d['activities']) for d in itinerary)} ativ | {len(bank_activities)} banco | {len(links)} dicas | {len(expenses)} gastos | {len(attachments)} anexos")
