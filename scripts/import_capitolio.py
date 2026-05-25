import openpyxl, json, os, re, sys
from import_utils import build_merge_map, make_activity_fn
sys.stdout.reconfigure(encoding='utf-8')

FOLDER = r'G:\Meu Drive\Viagens\2021-10 Capit' + chr(243) + 'lio'
XLSX   = FOLDER + r'\Viagem Capit' + chr(243) + 'lio.xlsx'
PFX    = 'capitolio-2021'

wb      = openpyxl.load_workbook(XLSX)
wb_data = openpyxl.load_workbook(XLSX, data_only=True)
ws      = wb['Roteiro']
merge_map = build_merge_map(ws)
act = make_activity_fn(ws, merge_map, first_col=4)

DATES = ["2021-10-23","2021-10-24","2021-10-25","2021-10-26","2021-10-27",
         "2021-10-28","2021-10-29","2021-10-30","2021-10-31","2021-11-01","2021-11-02"]
SUMMARIES = ["Ida","Cap" + chr(237) + "t" + chr(243) + "lio 1","Cap" + chr(237) + "t" + chr(243) + "lio 2",
             "Cap" + chr(237) + "t" + chr(243) + "lio 3","Cap" + chr(237) + "t" + chr(243) + "lio 4",
             "Cap" + chr(237) + "t" + chr(243) + "lio 5","Caxambu",
             "Penedo 1","Penedo 2","Penedo 3","Volta"]

DAYS_DEF = [
    (3,  0, [(4,'a01','Transporte'),(6,'a02','Refeicao'),(8,'a03','Transporte'),(10,'a04','Hospedagem')]),
    (4,  1, [(4,'a05','Passeio'),(6,'a06','Refeicao'),(9,'a07','Passeio'),(10,'a08','Hospedagem')]),
    (5,  2, [(4,'a09','Passeio'),(6,'a10','Refeicao'),(7,'a11','Passeio'),(10,'a12','Hospedagem')]),
    (6,  3, [(4,'a13','Passeio'),(6,'a14','Refeicao'),(7,'a15','Passeio'),(10,'a16','Hospedagem')]),
    (7,  4, [(4,'a17','Passeio'),(6,'a18','Refeicao'),(7,'a19','Passeio'),(10,'a20','Hospedagem')]),
    (8,  5, [(4,'a21','Passeio'),(6,'a22','Refeicao'),(10,'a23','Hospedagem')]),
    (9,  6, [(4,'a24','Transporte'),(6,'a25','Refeicao'),(7,'a26','Passeio'),(10,'a27','Hospedagem')]),
    (10, 7, [(4,'a28','Transporte'),(6,'a29','Passeio'),(10,'a30','Hospedagem')]),
    (11, 8, [(4,'a31','Passeio'),(10,'a32','Hospedagem')]),
    (12, 9, [(4,'a33','Passeio'),(10,'a34','Hospedagem')]),
    (13,10, [(4,'a35','Transporte')]),
]

itinerary = []
for (row, di, cols) in DAYS_DEF:
    activities = []
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = act(f"{PFX}-{sid}", row, col, atype)
        if a: activities.append(a)
    itinerary.append({"id":f"{PFX}-d{di+1:02d}","date":DATES[di],
                      "title":f"Dia {di+1}","summary":SUMMARIES[di],"activities":activities})

BANK_DEF = [
    (15, 0, [(4,'b01','Passeio'),(8,'b02','Passeio')]),
    (16, 1, [(4,'b03','Passeio'),(7,'b04','Passeio')]),
    (17, 2, [(4,'b05','Passeio'),(5,'b06','Passeio'),(6,'b07','Passeio'),(7,'b08','Passeio'),(8,'b09','Passeio')]),
]
bank_activities = []
for (row, br, cols) in BANK_DEF:
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = act(f"{PFX}-{sid}", row, col, atype, bankrow=br)
        if a: bank_activities.append(a)

ws_d = wb['Dicas']
links, seen_urls, lid = [], set(), 1
maps_url = ""
for row in ws_d.iter_rows(min_row=2):
    tv = row[0].value; uv = row[1].value if len(row)>1 else None
    if not tv or not uv: continue
    url = re.sub(r'\+[A-Z]\d+$','',str(uv).strip())
    title = str(tv).strip()
    if not url.startswith('http'): continue
    if title == 'Mapa': maps_url = url; continue
    if url in seen_urls: continue
    seen_urls.add(url)
    links.append({"id":f"{PFX}-l{lid:02d}","title":title,"url":url}); lid += 1

def parse_gastos_brl(ws_g, ws_g_data, pfx, pi, ti, ppli, qi, pagi):
    expenses = []; eid = 1
    for rr, rd in zip(ws_g.iter_rows(min_row=2), ws_g_data.iter_rows(min_row=2)):
        raw = list(rr); dat = list(rd)
        if str(raw[0].value or '').strip() not in ('>', chr(9658), chr(9654)): continue
        title = str(raw[1].value or '').strip()
        etype = str(raw[2].value or '').strip() or None
        obs1  = str(raw[3].value or '').strip()
        obs2  = str(raw[4].value or '').strip() if len(raw)>4 else ''
        obs3  = str(raw[5].value or '').strip() if len(raw)>5 else ''
        price_r = raw[pi].value; price_d = dat[pi].value
        price = float(price_r) if isinstance(price_r,(int,float)) else (float(price_d) if isinstance(price_d,(int,float)) else 0.0)
        taxes_r = raw[ti].value if ti is not None else None
        taxes = float(taxes_r) if isinstance(taxes_r,(int,float)) else 0.0
        ppl_r = raw[ppli].value; ppl_d = dat[ppli].value
        people = int(round(float(ppl_r))) if isinstance(ppl_r,(int,float)) else (int(round(float(ppl_d))) if isinstance(ppl_d,(int,float)) else 1)
        qty    = int(raw[qi].value or 1) if raw[qi].value else 1
        paid_d = dat[pagi].value; paid_r = raw[pagi].value
        paid = float(paid_d) if isinstance(paid_d,(int,float)) else (float(paid_r) if isinstance(paid_r,(int,float)) else 0.0)
        if not title: title = obs1 if obs1 and obs1!='-' else (obs2 if obs2 and obs2!='-' else 'Item')
        notes = ' . '.join(p for p in [obs1,obs2,obs3] if p and p not in ('-','')) or None
        if etype == 'Hotel': etype = 'Hospedagem'
        if etype in ('Atividades','Passeios e Ingressos'): etype = 'Passeios'
        if not etype:
            tl = title.lower()
            if any(w in tl for w in ['hotel','pousada','casa','voa']): etype='Hospedagem'
            elif any(w in tl for w in ['voo','aluguel','gasolina','traslado']): etype='Transporte'
            elif any(w in tl for w in ['ingresso','passeio','mirante','canyon','trilha','parque','pou']): etype='Passeios'
        e = {"id":f"{pfx}-e{eid:02d}","isActive":True,"title":title}
        if etype: e["type"] = etype
        if notes: e["notes"] = notes
        e.update({"price":round(price,2),"taxes":round(taxes,2),"people":people,"quantity":qty,
                  "currency":"BRL","exchangeRateToBase":1.0,"paidAmount":round(paid,2)})
        expenses.append(e); eid += 1
    return expenses

expenses = parse_gastos_brl(wb['Gastos'], wb_data['Gastos'], PFX, 6, None, 8, 9, 11)

files = sorted([f for f in os.listdir(FOLDER) if not f.startswith('.') and os.path.isfile(os.path.join(FOLDER,f))])
attachments = [{"id":f"{PFX}-att{i+1:02d}","file":f} for i,f in enumerate(files)]

title_str = "2021-10 Cap" + chr(237) + "t" + chr(243) + "lio"
trip = {
    "schemaVersion":1,"id":PFX,"title":title_str,
    "startDate":"2021-10-23","endDate":"2021-11-02",
    "baseCurrency":"BRL","people":2,"rateDecimalDigits":2,"myMapsUrl":maps_url,
    "itinerarySlotsPerDay":7,
    "itineraryVersions":[{"id":f"{PFX}-v1","name":"Vers" + chr(227) + "o 1","bankRows":3,
                          "itinerary":itinerary,"bankActivities":bank_activities}],
    "activeVersionId":f"{PFX}-v1",
    "tasks":[],"links":links,"expenses":expenses,"currencyRates":[],"attachments":attachments
}

out = os.path.join(FOLDER,'trip.json')
with open(out,'w',encoding='utf-8') as f: json.dump(trip,f,ensure_ascii=False,indent=2)
print(f"OK: {out}")
print(f"  {len(itinerary)} dias | {sum(len(d['activities']) for d in itinerary)} ativ | {len(bank_activities)} banco | {len(links)} dicas | {len(expenses)} gastos | {len(attachments)} anexos")
