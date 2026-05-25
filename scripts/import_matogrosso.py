import openpyxl, json, os, re, sys
from import_utils import build_merge_map, make_activity_fn
sys.stdout.reconfigure(encoding='utf-8')

XLSX   = r'G:\Meu Drive\Viagens\2023-07 Mato Grosso\Viagem Mato Grosso.xlsx'
FOLDER = r'G:\Meu Drive\Viagens\2023-07 Mato Grosso'
PFX    = 'matogrosso-2023'

wb      = openpyxl.load_workbook(XLSX)
wb_data = openpyxl.load_workbook(XLSX, data_only=True)
ws      = wb['Roteiro']
merge_map = build_merge_map(ws)
act = make_activity_fn(ws, merge_map, first_col=4)

DATES = ["2023-07-19","2023-07-20","2023-07-21","2023-07-22","2023-07-23",
         "2023-07-24","2023-07-25","2023-07-26","2023-07-27","2023-07-28"]
SUMMARIES = ["IDA","Cuiaba 1","Cuiaba 2","Pantanal 1","Pantanal 2",
             "Pantanal 3","Chapada 1","Chapada 2","Chapada 3","VOLTA"]

DAYS_DEF = [
    (3, 0, [(10,'a01','Transporte'),(12,'a02','Hospedagem')]),
    (4, 1, [(4,'a03','Passeio'),(8,'a04','Refeicao'),(10,'a05','Passeio'),(12,'a06','Hospedagem')]),
    (5, 2, [(4,'a07','Passeio'),(7,'a08','Refeicao'),(8,'a09','Transporte'),(12,'a10','Hospedagem')]),
    (6, 3, [(4,'a11','Passeio'),(7,'a12','Refeicao'),(9,'a13','Passeio'),(12,'a14','Hospedagem')]),
    (7, 4, [(4,'a15','Passeio'),(6,'a16','Passeio'),(7,'a17','Refeicao'),
            (8,'a18','Passeio'),(11,'a19','Passeio'),(12,'a20','Hospedagem')]),
    (8, 5, [(4,'a21','Passeio'),(7,'a22','Refeicao'),(8,'a23','Passeio'),
            (9,'a24','Passeio'),(12,'a25','Hospedagem')]),
    (9, 6, [(5,'a26','Transporte'),(7,'a27','Refeicao'),(8,'a28','Passeio'),
            (10,'a29','Passeio'),(12,'a30','Hospedagem')]),
    (10,7, [(4,'a31','Passeio'),(12,'a32','Hospedagem')]),
    (11,8, [(7,'a33','Refeicao'),(12,'a34','Hospedagem')]),
    (12,9, [(5,'a35','Transporte'),(6,'a36','Transporte')]),
]

itinerary = []
for (row, di, cols) in DAYS_DEF:
    activities = []
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = act(f"{PFX}-{sid}", row, col, atype)
        if a: activities.append(a)
    itinerary.append({"id":f"{PFX}-d{di+1:02d}","date":DATES[di],
                      "title":f"Dia {di+1}","summary":SUMMARIES[di],"activities":activities})

BANK_DEF = [
    (14, 0, [(4,'b01','Trilha'),(7,'b02','Trilha'),(9,'b03','Passeio'),(10,'b04','Passeio')]),
    (15, 1, [(4,'b05','Trilha'),(7,'b06','Trilha')]),
]
bank_activities = []
for (row, br, cols) in BANK_DEF:
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = act(f"{PFX}-{sid}", row, col, atype, bankrow=br)
        if a: bank_activities.append(a)

ws_d = wb['Dicas']
links, seen_urls, lid = [], set(), 1
maps_url = ""
for row in ws_d.iter_rows(min_row=2):
    tv = row[0].value; uv = row[1].value if len(row)>1 else None
    if not tv or not uv: continue
    url = re.sub(r'\+[A-Z]\d+$','',str(uv).strip())
    title = str(tv).strip()
    if title == 'Mapa': maps_url = url; continue
    if url in seen_urls: continue
    seen_urls.add(url)
    links.append({"id":f"{PFX}-l{lid:02d}","title":title,"url":url}); lid += 1

def parse_gastos_brl(ws_g, ws_g_data, pfx, pi, ti, ppli, qi, pagi):
    expenses = []; eid = 1
    for rr, rd in zip(ws_g.iter_rows(min_row=2), ws_g_data.iter_rows(min_row=2)):
        raw = list(rr); dat = list(rd)
        if str(raw[0].value or '').strip() not in (chr(9658), chr(9654)): continue
        title  = str(raw[1].value or '').strip()
        etype  = str(raw[2].value or '').strip() or None
        obs1   = str(raw[3].value or '').strip()
        obs2   = str(raw[4].value or '').strip() if len(raw)>4 else ''
        obs3   = str(raw[5].value or '').strip() if len(raw)>5 else ''
        price_r = raw[pi].value; price_d = dat[pi].value
        price = float(price_r) if isinstance(price_r,(int,float)) else (float(price_d) if isinstance(price_d,(int,float)) else 0.0)
        taxes_r = raw[ti].value if ti is not None else None
        taxes = float(taxes_r) if isinstance(taxes_r,(int,float)) else 0.0
        people = int(round(float(raw[ppli].value))) if isinstance(raw[ppli].value,(int,float)) else 1
        qty    = int(raw[qi].value or 1) if raw[qi].value else 1
        paid_d = dat[pagi].value; paid_r = raw[pagi].value
        paid = float(paid_d) if isinstance(paid_d,(int,float)) else (float(paid_r) if isinstance(paid_r,(int,float)) else 0.0)
        if not title: title = obs1 if obs1 and obs1!='-' else (obs2 if obs2 and obs2!='-' else 'Item')
        notes = ' . '.join(p for p in [obs1,obs2,obs3] if p and p not in ('-','')) or None
        if etype == 'Hotel': etype = 'Hospedagem'
        if etype == 'Atividades': etype = 'Passeios'
        if not etype:
            tl = title.lower()
            if any(w in tl for w in ['hotel','pousada','glamping','hostal']): etype='Hospedagem'
            elif any(w in tl for w in ['voo','aluguel','traslado','transfer']): etype='Transporte'
            elif any(w in tl for w in ['ingresso','excursao','guia','trilha','cidade','passeio']): etype='Passeios'
        e = {"id":f"{pfx}-e{eid:02d}","isActive":True,"title":title}
        if etype: e["type"] = etype
        if notes: e["notes"] = notes
        e.update({"price":round(price,2),"taxes":round(taxes,2),"people":people,"quantity":qty,
                  "currency":"BRL","exchangeRateToBase":1.0,"paidAmount":round(paid,2)})
        expenses.append(e); eid += 1
    return expenses

expenses = parse_gastos_brl(wb['Gastos'], wb_data['Gastos'], PFX, 7, 8, 9, 10, 12)

files = sorted([f for f in os.listdir(FOLDER) if not f.startswith('.') and os.path.isfile(os.path.join(FOLDER,f))])
attachments = [{"id":f"{PFX}-att{i+1:02d}","file":f} for i,f in enumerate(files)]

trip = {
    "schemaVersion":1,"id":PFX,"title":"2023-07 Mato Grosso",
    "startDate":"2023-07-19","endDate":"2023-07-28",
    "baseCurrency":"BRL","people":4,"rateDecimalDigits":2,"myMapsUrl":maps_url,
    "itinerarySlotsPerDay":9,
    "itineraryVersions":[{"id":f"{PFX}-v1","name":"Versao 1","bankRows":2,
                          "itinerary":itinerary,"bankActivities":bank_activities}],
    "activeVersionId":f"{PFX}-v1",
    "tasks":[],"links":links,"expenses":expenses,"currencyRates":[],"attachments":attachments
}

out = os.path.join(FOLDER,'trip.json')
with open(out,'w',encoding='utf-8') as f: json.dump(trip,f,ensure_ascii=False,indent=2)
print(f"OK: {out}")
print(f"  {len(itinerary)} dias | {sum(len(d['activities']) for d in itinerary)} ativ | {len(bank_activities)} banco | {len(links)} dicas | {len(expenses)} gastos | {len(attachments)} anexos")
