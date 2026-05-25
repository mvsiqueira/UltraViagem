import openpyxl, json, os, re, sys
from import_utils import cell_color, parse_text, build_merge_map, make_activity_fn, parse_expenses
sys.stdout.reconfigure(encoding='utf-8')

XLSX   = r'G:\Meu Drive\Viagens\2025-11 Carretera Austral\Viagem Carretera Austral.xlsx'
FOLDER = r'G:\Meu Drive\Viagens\2025-11 Carretera Austral'
PFX    = 'carretera-2025'

wb      = openpyxl.load_workbook(XLSX)
wb_data = openpyxl.load_workbook(XLSX, data_only=True)
ws      = wb['Roteiro']
merge_map = build_merge_map(ws)
act = make_activity_fn(ws, merge_map, first_col=5)

# Cols E(5)-O(15) = slots 0-10 → 11 slots
# Col B = data, C = dia da semana, D = título do dia
# Rows 3-18 = 16 dias; Opcionais cols R+ → ignorados

DATES = [
    "2025-11-27","2025-11-28","2025-11-29","2025-11-30",
    "2025-12-01","2025-12-02","2025-12-03","2025-12-04",
    "2025-12-05","2025-12-06","2025-12-07","2025-12-08",
    "2025-12-09","2025-12-10","2025-12-11","2025-12-12",
]
TITLES = ["Ida"] + [f"Dia {i}" for i in range(1,16)]
SUMMARIES = [
    "IDA","Coyhaique","Puerto Río Tranquilo","Cochrane",
    "Cochrane","Puerto Río Tranquilo","Coyhaique","Villa Cerro Castillo",
    "Coyhaique","Puyuhuapi","Futaleufú","Chaitén",
    "La Junta","Coyhaique","Santiago","VOLTA",
]

# (excel_row, day_idx, [(col, id_suffix, type)])
DAYS_DEF = [
    # Ida
    (3, 0, [(9,'a01','Transporte'),(15,'a02','Hospedagem')]),
    # Dia 1
    (4, 1, [(5,'a03','Passeio'),(8,'a04','Refeição'),(9,'a05','Transporte'),
            (12,'a06','Transporte'),(15,'a07','Hospedagem')]),
    # Dia 2
    (5, 2, [(5,'a08','Transporte'),(6,'a09','Transporte'),(7,'a10','Transporte'),
            (8,'a11','Refeição'),(9,'a12','Transporte'),(11,'a13','Passeio'),
            (15,'a14','Hospedagem')]),
    # Dia 3
    (6, 3, [(5,'a15','Passeio'),(8,'a16','Refeição'),(9,'a17','Passeio'),
            (11,'a18','Transporte'),(12,'a19','Transporte'),(13,'a20','Passeio'),
            (15,'a21','Hospedagem')]),
    # Dia 4
    (7, 4, [(5,'a22','Transporte'),(7,'a23','Passeio'),(12,'a24','Transporte'),
            (15,'a25','Hospedagem')]),
    # Dia 5
    (8, 5, [(5,'a26','Trilha'),(8,'a27','Refeição'),(9,'a28','Passeio'),
            (12,'a29','Transporte'),(13,'a30','Transporte'),(15,'a31','Hospedagem')]),
    # Dia 6
    (9, 6, [(5,'a32','Trilha'),(8,'a33','Refeição'),(9,'a34','Passeio'),
            (10,'a35','Transporte'),(12,'a36','Transporte'),(13,'a37','Transporte'),
            (15,'a38','Hospedagem')]),
    # Dia 7
    (10,7, [(5,'a39','Transporte'),(6,'a40','Transporte'),(8,'a41','Refeição'),
            (9,'a42','Passeio'),(10,'a43','Passeio'),(11,'a44','Transporte'),
            (12,'a45','Transporte'),(13,'a46','Passeio'),(15,'a47','Hospedagem')]),
    # Dia 8
    (11,8, [(5,'a48','Trilha'),(11,'a49','Transporte'),(12,'a50','Transporte'),
            (13,'a51','Passeio'),(15,'a52','Hospedagem')]),
    # Dia 9
    (12,9, [(5,'a53','Transporte'),(6,'a54','Transporte'),(7,'a55','Transporte'),
            (8,'a56','Refeição'),(9,'a57','Trilha'),(12,'a58','Passeio'),
            (15,'a59','Hospedagem')]),
    # Dia 10
    (13,10,[(5,'a60','Transporte'),(6,'a61','Transporte'),(7,'a62','Transporte'),
            (8,'a63','Refeição'),(9,'a64','Passeio'),(15,'a65','Hospedagem')]),
    # Dia 11
    (14,11,[(5,'a66','Transporte'),(6,'a67','Transporte'),(7,'a68','Transporte'),
            (8,'a69','Refeição'),(9,'a70','Trilha'),(12,'a71','Passeio'),
            (15,'a72','Hospedagem')]),
    # Dia 12
    (15,12,[(5,'a73','Transporte'),(6,'a74','Trilha'),(8,'a75','Refeição'),
            (9,'a76','Transporte'),(10,'a77','Transporte'),(11,'a78','Passeio'),
            (15,'a79','Hospedagem')]),
    # Dia 13
    (16,13,[(5,'a80','Transporte'),(6,'a81','Trilha'),(8,'a82','Refeição'),
            (9,'a83','Transporte'),(10,'a84','Transporte'),(11,'a85','Transporte'),
            (12,'a86','Transporte'),(13,'a87','Transporte'),(15,'a88','Hospedagem')]),
    # Dia 14
    (17,14,[(6,'a89','Transporte'),(8,'a90','Refeição'),(9,'a91','Transporte'),
            (11,'a92','Transporte'),(15,'a93','Hospedagem')]),
    # Dia 15
    (18,15,[(7,'a94','Transporte')]),
]

itinerary = []
for (row, di, cols) in DAYS_DEF:
    activities = []
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = act(f"{PFX}-{sid}", row, col, atype)
        if a: activities.append(a)
    itinerary.append({"id":f"{PFX}-d{di+1:02d}","date":DATES[di],
                      "title":TITLES[di],"summary":SUMMARIES[di],"activities":activities})

# Banco: opcionais mapeados como 2 bank rows
# Row 5 Opcionais: Mirador Glaciar Exploradores (col S5:U5 → ignorar colunas R+)
# Vou usar apenas atividades claramente identificadas nas colunas R-X
# Por ora banco vazio (estrutura incompatível com slot system dos opcionais)
bank_activities = []

# Dicas + mapa
ws_d = wb['Dicas']
links, seen_urls, lid = [], set(), 1
maps_url = ""
for row in ws_d.iter_rows(min_row=2):
    tv = row[0].value; uv = row[1].value if len(row)>1 else None
    if not tv or not uv: continue
    url = re.sub(r'\+[A-Z]\d+$','',str(uv).strip())
    title = str(tv).strip()
    if title == 'Mapa':
        maps_url = url
        continue  # não inclui o mapa na lista de dicas
    if url in seen_urls: continue
    seen_urls.add(url)
    links.append({"id":f"{PFX}-l{lid:02d}","title":title,"url":url})
    lid += 1

# Gastos — Carretera usa CLP/USD (não Peso/Dólar)
moeda_keys = {'CLP':'CLP','USD':'USD','Peso':'CLP','Dólar':'USD','Dolar':'USD'}
expenses = parse_expenses(wb['Gastos'], wb_data['Gastos'], PFX, moeda_keys)

files = sorted([f for f in os.listdir(FOLDER)
                if not f.startswith('.') and os.path.isfile(os.path.join(FOLDER,f))])
attachments = [{"id":f"{PFX}-att{i+1:02d}","file":f} for i,f in enumerate(files)]

trip = {
    "schemaVersion":1,"id":PFX,"title":"2025-11 Carretera Austral",
    "startDate":"2025-11-27","endDate":"2025-12-12",
    "baseCurrency":"BRL","people":2,"rateDecimalDigits":2,"myMapsUrl":maps_url,
    "itinerarySlotsPerDay":11,
    "itineraryVersions":[{"id":f"{PFX}-v1","name":"Versão 1","bankRows":0,
                          "itinerary":itinerary,"bankActivities":bank_activities}],
    "activeVersionId":f"{PFX}-v1",
    "tasks":[],"links":links,"expenses":expenses,"currencyRates":[],"attachments":attachments
}

out = os.path.join(FOLDER,'trip.json')
with open(out,'w',encoding='utf-8') as f: json.dump(trip,f,ensure_ascii=False,indent=2)
print(f"OK: {out}")
print(f"  {len(itinerary)} dias | {sum(len(d['activities']) for d in itinerary)} ativ | {len(bank_activities)} banco | {len(links)} dicas | {len(expenses)} gastos | {len(attachments)} anexos")
print()
print("Atividades:")
for d in itinerary:
    acts_str = '  '.join(f"[{a['startSlot']}+{a['durationSlots']}]{a['title'][:20]}" for a in d['activities'])
    print(f"  {d['date']} {d['summary']:25} {acts_str}")
print()
print("Gastos:")
for e in expenses:
    print(f"  {e['id']}: {e['title'][:38]:38} {e['currency']} p={e['price']} ppl={e['people']} qty={e['quantity']} paid={e['paidAmount']}")
