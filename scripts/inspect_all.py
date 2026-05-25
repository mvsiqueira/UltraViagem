import openpyxl, sys, os
sys.stdout.reconfigure(encoding='utf-8')

TRIPS = [
    (r'G:\Meu Drive\Viagens\2023-02 Buenos Aires',           r'Viagem Buenos Aires1.xlsx'),
    (r'G:\Meu Drive\Viagens\2023-04 Chapada Diamantina',     r'Viagem Chapada Diamantina2.xlsx'),
    (r'G:\Meu Drive\Viagens\2023-07 Mato Grosso',            r'Viagem Mato Grosso.xlsx'),
    (r'G:\Meu Drive\Viagens\2023-09 São Paulo',              r'Viagem São Paulo.xlsx'),
    (r'G:\Meu Drive\Viagens\2023-11 Portugal',               r'Viagem Portugal.xlsx'),
    (r'G:\Meu Drive\Viagens\2024-04 Salar de Uyuni',         r'Viagem Salar de Uyuni.xlsx'),
    (r'G:\Meu Drive\Viagens\2024-07 Porto Seguro',           r'Viagem Porto Seguro.xlsx'),
    (r'G:\Meu Drive\Viagens\2024-11 Emirados Árabes Unidos', r'Viagem Emirados Arabes.xlsx'),
    (r'G:\Meu Drive\Viagens\2024-11 Orlando',                r'Viagem Disney 2024.xlsx'),
]

for folder, fname in TRIPS:
    path = os.path.join(folder, fname)
    print(f"\n{'='*70}")
    print(f"### {os.path.basename(folder)} ###")
    print(f"{'='*70}")
    wb = openpyxl.load_workbook(path)
    print(f"Abas: {wb.sheetnames}")

    ws = wb['Roteiro']
    print(f"Dimensoes: {ws.dimensions}  max_row={ws.max_row} max_col={ws.max_column}")

    print("MERGES:")
    for mc in sorted(ws.merged_cells.ranges, key=lambda m:(m.min_row,m.min_col)):
        real = ws.cell(row=mc.min_row, column=mc.min_col)
        v = str(real.value or '').replace('\n','|')[:55]
        fg = real.fill.fgColor
        color = fg.rgb if fg.type=='rgb' else (f"theme{fg.theme}_t{fg.tint:.2f}" if fg.type=='theme' else '?')
        print(f"  {str(mc):12} r{mc.min_row}c{mc.min_col}:r{mc.max_row}c{mc.max_col}  {color}  {v!r}")

    print("CELULAS:")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.value is None: continue
            v = str(cell.value).replace('\n','|')
            fg = cell.fill.fgColor
            color = fg.rgb if fg.type=='rgb' else (f"theme{fg.theme}_t{fg.tint:.2f}" if fg.type=='theme' else '?')
            print(f"  [{cell.row},{cell.column}]{cell.coordinate}: {v!r}  {color}")

    if 'Dicas' in wb.sheetnames:
        print("DICAS:")
        for row in wb['Dicas'].iter_rows(min_row=2):
            tv = row[0].value; uv = row[1].value if len(row)>1 else None
            if tv or uv:
                print(f"  {str(tv or '')[:35]!r}  {str(uv or '')[:60]!r}")

    if 'Gastos' in wb.sheetnames:
        print("GASTOS (header + marcados):")
        ws_g = wb['Gastos']
        for i, row in enumerate(ws_g.iter_rows(min_row=1), 1):
            vals = [c.value for c in row]
            marker = str(vals[0] or '').strip()
            if i == 1 or marker in ('►','▶','>'):
                safe = [str(v or '')[:35] for v in vals]
                print(f"  row{i}: {safe}")
