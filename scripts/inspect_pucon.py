import openpyxl

wb = openpyxl.load_workbook(r'G:\Meu Drive\Viagens\2025-07 Pucón\Viagem Pucón.xlsx')

# ── ROTEIRO ──────────────────────────────────────────────────────────────────
ws = wb['Roteiro']
print(f"Dimensoes: {ws.dimensions}  max_row={ws.max_row} max_col={ws.max_column}")
print()

print("=== MERGED CELLS ===")
for mc in ws.merged_cells.ranges:
    real = ws.cell(row=mc.min_row, column=mc.min_col)
    v = str(real.value or '').replace('\n', ' | ')[:50]
    print(f"  {mc} -> [{mc.min_row},{mc.min_col}:{mc.max_row},{mc.max_col}]  '{v}'")

print()
print("=== CELULAS PREENCHIDAS ===")
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for cell in row:
        if cell.value is None:
            continue
        fg = cell.fill.fgColor
        if fg.type == 'rgb':
            color = fg.rgb
        elif fg.type == 'theme':
            color = f"theme{fg.theme}_tint{fg.tint:.3f}"
        else:
            color = fg.type
        v = str(cell.value).replace('\n', ' | ')[:60]
        print(f"  [{cell.row},{cell.column}] {cell.coordinate}: {repr(v)}  fill={cell.fill.fill_type} color={color}")

# ── DICAS ─────────────────────────────────────────────────────────────────────
print()
print("=== DICAS ===")
ws_d = wb['Dicas']
for row in ws_d.iter_rows(min_row=1, max_row=ws_d.max_row):
    vals = [str(c.value or '').strip() for c in row]
    if any(vals):
        print("  ", vals)

# ── GASTOS ────────────────────────────────────────────────────────────────────
print()
print("=== GASTOS (cabecalho + linhas marcadas) ===")
ws_g = wb['Gastos']
for i, row in enumerate(ws_g.iter_rows(min_row=1, max_row=ws_g.max_row), start=1):
    vals = [str(c.value or '').strip() for c in row]
    marker = vals[0] if vals else ''
    if i == 1 or marker in ('►', '▶'):
        print(f"  row{i}: {vals}")
