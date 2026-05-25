import openpyxl, json, os, re, sys
from import_utils import build_merge_map, make_activity_fn
sys.stdout.reconfigure(encoding='utf-8')

XLSX   = r'G:\Meu Drive\Viagens\2024-11 Orlando\Viagem Disney 2024.xlsx'
FOLDER = r'G:\Meu Drive\Viagens\2024-11 Orlando'
PFX    = 'orlando-2024'

wb      = openpyxl.load_workbook(XLSX)
wb_data = openpyxl.load_workbook(XLSX, data_only=True)
ws      = wb['Roteiro']
merge_map = build_merge_map(ws)
act = make_activity_fn(ws, merge_map, first_col=4)

DATES = ["2024-11-25","2024-11-26","2024-11-27","2024-11-28","2024-11-29",
         "2024-11-30","2024-12-01","2024-12-02","2024-12-03","2024-12-04"]
SUMMARIES = ["IDA","Walmart","Christmas Party","Disney Springs","Animal Kingdom",
             "CityWalk","Outlet","Livre","Volta","Volta 2"]

DAYS_DEF = [
    (3,  0, [(4,'a01','Transporte'),(12,'a02','Hospedagem')]),
    (4,  1, [(4,'a03','Passeio'),(8,'a04','Passeio'),(12,'a05','Hospedagem')]),
    (5,  2, [(4,'a06','Passeio'),(8,'a07','Passeio'),(12,'a08','Hospedagem')]),
    (6,  3, [(8,'a09','Passeio'),(12,'a10','Hospedagem')]),
    (7,  4, [(4,'a11','Passeio'),(5,'a12','Passeio'),(12,'a13','Hospedagem')]),
    (8,  5, [(4,'a14','Passeio'),(6,'a15','Passeio'),(9,'a16','Passeio'),(12,'a17','Hospedagem')]),
    (9,  6, [(4,'a18','Passeio'),(12,'a19','Hospedagem')]),
    (10, 7, [(12,'a20','Hospedagem')]),
    (11, 8, [(8,'a21','Transporte')]),
    (12, 9, [(4,'a22','Transporte')]),
]

itinerary = []
for (row, di, cols) in DAYS_DEF:
    activities = []
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = act(f"{PFX}-{sid}", row, col, atype)
        if a: activities.append(a)
    itinerary.append({"id":f"{PFX}-d{di+1:02d}","date":DATES[di],
                      "title":f"Dia {di+1}","summary":SUMMARIES[di],"activities":activities})

BANK_DEF = [
    (14, 0, [(4,'b01','Passeio'),(6,'b02','Passeio'),(8,'b03','Passeio'),(10,'b04','Passeio')]),
    (15, 1, [(4,'b05','Passeio'),(6,'b06','Passeio'),(8,'b07','Passeio'),(10,'b08','Passeio')]),
    (16, 2, [(4,'b09','Passeio'),(6,'b10','Passeio'),(8,'b11','Passeio')]),
]
bank_activities = []
for (row, br, cols) in BANK_DEF:
    for (col, sid, atype) in cols:
        mc = merge_map.get((row, col))
        if mc and mc.min_col != col: continue
        a = act(f"{PFX}-{sid}", row, col, atype, bankrow=br)
        if a: bank_activities.append(a)

ws_d = wb['Dicas']
links, seen_urls, lid = [], set(), 1
maps_url = ""
for row in ws_d.iter_rows(min_row=2):
    tv = row[0].value; uv = row[1].value if len(row)>1 else None
    if not tv or not uv: continue
    url = re.sub(r'\+[A-Z]\d+$','',str(uv).strip())
    title = str(tv).strip()
    if title == 'Mapa': maps_url = url; continue
    if url in seen_urls: continue
    seen_urls.add(url)
    links.append({"id":f"{PFX}-l{lid:02d}","title":title,"url":url}); lid += 1

# Gastos Orlando: standard cols; moeda Dolar->USD or hardcoded numeric rate ~5.5->USD
def parse_gastos_orlando(ws_g, ws_g_data, pfx):
    expenses = []; eid = 1
    for rr, rd in zip(ws_g.iter_rows(min_row=2), ws_g_data.iter_rows(min_row=2)):
        raw = list(rr); dat = list(rd)
        if str(raw[0].value or '').strip() not in (chr(9658), chr(9654)): continue
        title  = str(raw[1].value or '').strip()
        etype  = str(raw[2].value or '').strip() or None
        obs1   = str(raw[3].value or '').strip()
        obs2   = str(raw[4].value or '').strip() if len(raw)>4 else ''
        obs3   = str(raw[5].value or '').strip() if len(raw)>5 else ''
        price_r = raw[8].value; price_d = dat[8].value
        price = float(price_r) if isinstance(price_r,(int,float)) else (float(price_d) if isinstance(price_d,(int,float)) else 0.0)
        taxes_r = raw[9].value
        taxes = float(taxes_r) if isinstance(taxes_r,(int,float)) else 0.0
        people = int(round(float(raw[10].value))) if isinstance(raw[10].value,(int,float)) else 1
        qty    = int(raw[11].value or 1) if raw[11].value else 1
        moeda_raw  = str(raw[13].value or '1').strip()
        moeda_calc = dat[13].value
        currency, rate = 'BRL', 1.0
        if 'Dolar' in moeda_raw or chr(243) in moeda_raw:  # Dólar
            currency = 'USD'; rate = float(moeda_calc) if isinstance(moeda_calc,(int,float)) else 0.0
        elif moeda_raw == '1':
            currency = 'BRL'; rate = 1.0
        else:
            try:
                r = float(moeda_raw)
                if r > 3.0: currency, rate = 'USD', r
            except ValueError: pass
        paid_d = dat[15].value; paid_r = raw[15].value
        paid = float(paid_d) if isinstance(paid_d,(int,float)) else (float(paid_r) if isinstance(paid_r,(int,float)) else 0.0)
        if not title: title = obs1 if obs1 and obs1!='-' else (obs2 if obs2 and obs2!='-' else 'Item')
        notes = ' . '.join(p for p in [obs1,obs2,obs3] if p and p not in ('-','')) or None
        if etype == 'Hotel': etype = 'Hospedagem'
        if etype == 'Atividades': etype = 'Passeios'
        if not etype:
            tl = title.lower()
            if any(w in tl for w in ['hotel','blue tree','hostel']): etype='Hospedagem'
            elif any(w in tl for w in ['voo','aluguel','estacionamento','mala','assento']): etype='Transporte'
            elif any(w in tl for w in ['ingresso','disney','lightning','universal']): etype='Passeios'
        e = {"id":f"{pfx}-e{eid:02d}","isActive":True,"title":title}
        if etype: e["type"] = etype
        if notes: e["notes"] = notes
        e.update({"price":round(price,2),"taxes":round(taxes,2),"people":people,"quantity":qty,
                  "currency":currency,"exchangeRateToBase":round(rate,6),"paidAmount":round(paid,2)})
        expenses.append(e); eid += 1
    return expenses

expenses = parse_gastos_orlando(wb['Gastos'], wb_data['Gastos'], PFX)

files = sorted([f for f in os.listdir(FOLDER) if not f.startswith('.') and os.path.isfile(os.path.join(FOLDER,f))])
attachments = [{"id":f"{PFX}-att{i+1:02d}","file":f} for i,f in enumerate(files)]

trip = {
    "schemaVersion":1,"id":PFX,"title":"2024-11 Orlando",
    "startDate":"2024-11-25","endDate":"2024-12-04",
    "baseCurrency":"BRL","people":3,"rateDecimalDigits":2,"myMapsUrl":maps_url,
    "itinerarySlotsPerDay":9,
    "itineraryVersions":[{"id":f"{PFX}-v1","name":"Versao 1","bankRows":3,
                          "itinerary":itinerary,"bankActivities":bank_activities}],
    "activeVersionId":f"{PFX}-v1",
    "tasks":[],"links":links,"expenses":expenses,"currencyRates":[],"attachments":attachments
}

out = os.path.join(FOLDER,'trip.json')
with open(out,'w',encoding='utf-8') as f: json.dump(trip,f,ensure_ascii=False,indent=2)
print(f"OK: {out}")
print(f"  {len(itinerary)} dias | {sum(len(d['activities']) for d in itinerary)} ativ | {len(bank_activities)} banco | {len(links)} dicas | {len(expenses)} gastos | {len(attachments)} anexos")
